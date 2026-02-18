"""
ACENCIA ATLAS - Dokumentenarchiv

Ansicht fuer alle Dokumente mit Upload/Download-Funktionen, PDF-Vorschau
und KI-basierter automatischer Benennung.
"""

from typing import Optional, List, Tuple
from datetime import datetime
import tempfile
import os
import logging

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QHeaderView, QPushButton, QLabel, QComboBox, QLineEdit,
    QFileDialog, QMessageBox, QMenu, QProgressDialog, QFrame,
    QSplitter, QGroupBox, QFormLayout, QDialog, QToolBar, QApplication
)
from PySide6.QtCore import Qt, Signal, QThread, QUrl, QTimer
from PySide6.QtGui import QAction, QFont, QColor

logger = logging.getLogger(__name__)

# PDF-Viewer: Versuche QPdfView zu importieren (Qt6 native PDF)
HAS_PDF_VIEW = False
HAS_WEBENGINE = False

try:
    from PySide6.QtPdfWidgets import QPdfView
    from PySide6.QtPdf import QPdfDocument
    HAS_PDF_VIEW = True
except ImportError:
    pass

# Fallback: QWebEngineView (braucht PDF.js workaround)
if not HAS_PDF_VIEW:
    try:
        from PySide6.QtWebEngineWidgets import QWebEngineView
        HAS_WEBENGINE = True
    except ImportError:
        pass

from api.client import APIClient
from api.documents import DocumentsAPI, Document


def format_date_german(date_str: str) -> str:
    """Konvertiert ISO-Datum/Datetime ins deutsche Format (DD.MM.YYYY).
    
    Unterstuetzt: 'YYYY-MM-DD', 'YYYY-MM-DDTHH:MM:SS', 'YYYY-MM-DD HH:MM:SS'
    """
    if not date_str:
        return ""
    try:
        # Zuerst nur den Datumsteil extrahieren (vor T oder Leerzeichen)
        date_part = date_str.strip()
        if 'T' in date_part:
            date_part = date_part.split('T')[0]
        elif ' ' in date_part:
            date_part = date_part.split(' ')[0]
        parts = date_part.split('-')
        if len(parts) == 3:
            year, month, day = parts
            return f"{day}.{month}.{year}"
    except (ValueError, IndexError):
        pass
    return date_str


class DocumentLoadWorker(QThread):
    """Worker zum Laden der Dokumente."""
    finished = Signal(list)
    error = Signal(str)
    
    def __init__(self, docs_api: DocumentsAPI, filters: dict):
        super().__init__()
        self.docs_api = docs_api
        self.filters = filters
    
    def run(self):
        try:
            docs = self.docs_api.list_documents(**self.filters)
            self.finished.emit(docs)
        except Exception as e:
            self.error.emit(str(e))


class UploadWorker(QThread):
    """Worker zum Hochladen von Dateien."""
    finished = Signal(object)  # Document or None
    error = Signal(str)
    progress = Signal(str)
    
    def __init__(self, docs_api: DocumentsAPI, file_path: str, source_type: str):
        super().__init__()
        self.docs_api = docs_api
        self.file_path = file_path
        self.source_type = source_type
    
    def run(self):
        try:
            self.progress.emit("Lade hoch...")
            doc = self.docs_api.upload(self.file_path, self.source_type)
            self.finished.emit(doc)
        except Exception as e:
            self.error.emit(str(e))


class AIRenameWorker(QThread):
    """
    Worker fuer KI-basierte PDF-Benennung.
    
    Verarbeitet PDFs im Hintergrund mit zweistufigem KI-System:
    1. Laedt PDF herunter
    2. Stufe 1 (Triage): Schnelle Kategorisierung mit GPT-4o-mini
    3. Stufe 2 (Detail): Detailanalyse mit GPT-4o nur bei Bedarf
    4. Umbenennung auf Server
    """
    # Signale
    finished = Signal(list)  # Liste von (doc_id, success, new_name_or_error)
    progress = Signal(int, int, str)  # current, total, current_filename
    single_finished = Signal(int, bool, str)  # doc_id, success, new_name_or_error
    error = Signal(str)
    
    def __init__(self, api_client, docs_api: DocumentsAPI, documents: List):
        super().__init__()
        self.api_client = api_client
        self.docs_api = docs_api
        self.documents = documents
        self._cancelled = False
    
    def cancel(self):
        """Abbrechen der Verarbeitung."""
        self._cancelled = True
    
    def run(self):
        from api.openrouter import OpenRouterClient, DocumentClassification
        
        results = []
        total = len(self.documents)
        
        try:
            # OpenRouter Client initialisieren
            openrouter = OpenRouterClient(self.api_client)
            
            for i, doc in enumerate(self.documents):
                if self._cancelled:
                    logger.info("KI-Benennung abgebrochen")
                    break
                
                self.progress.emit(i + 1, total, doc.original_filename)
                logger.info(f"Verarbeite {i+1}/{total}: {doc.original_filename}")
                
                # Bereits umbenannt? Skip.
                if doc.ai_renamed:
                    logger.info(f"Ueberspringe bereits umbenanntes Dokument: {doc.original_filename}")
                    results.append((doc.id, True, doc.original_filename))
                    self.single_finished.emit(doc.id, True, doc.original_filename)
                    continue
                
                # Nicht PDF? Skip.
                if not doc.is_pdf:
                    logger.info(f"Ueberspringe Nicht-PDF: {doc.original_filename}")
                    results.append((doc.id, True, doc.original_filename))
                    self.single_finished.emit(doc.id, True, doc.original_filename)
                    continue
                
                try:
                    # PDF herunterladen (filename_override spart extra API-Call)
                    temp_dir = tempfile.mkdtemp(prefix='bipro_ai_')
                    pdf_path = self.docs_api.download(doc.id, temp_dir, filename_override=doc.original_filename)
                    
                    if not pdf_path or not os.path.exists(pdf_path):
                        raise Exception("Download fehlgeschlagen")
                    
                    # Zweistufige KI-Klassifikation (Triage -> Detail bei Bedarf)
                    classification = openrouter.classify_pdf_smart(pdf_path)
                    logger.info(f"Klassifikation: {classification.target_box} ({classification.confidence})")
                    
                    # Neuen Dateinamen generieren
                    original_ext = os.path.splitext(doc.original_filename)[1] or '.pdf'
                    new_filename = classification.generate_filename(original_ext)
                    
                    # Auf Server umbenennen
                    success = self.docs_api.rename_document(doc.id, new_filename, mark_ai_renamed=True)
                    
                    if success:
                        logger.info(f"Umbenannt: {doc.original_filename} -> {new_filename}")
                        results.append((doc.id, True, new_filename))
                        self.single_finished.emit(doc.id, True, new_filename)
                    else:
                        raise Exception("Server-Update fehlgeschlagen")
                    
                    # Temporaere Dateien aufraeumen
                    try:
                        os.unlink(pdf_path)
                        os.rmdir(temp_dir)
                    except OSError:
                        pass
                    
                except Exception as e:
                    error_msg = str(e)
                    logger.error(f"Fehler bei {doc.original_filename}: {error_msg}")
                    
                    # Fehler in DB speichern
                    try:
                        self.docs_api.update(doc.id, ai_processing_error=error_msg[:500])
                    except Exception:
                        pass
                    
                    results.append((doc.id, False, error_msg))
                    self.single_finished.emit(doc.id, False, error_msg)
            
            self.finished.emit(results)
            
        except Exception as e:
            logger.error(f"AIRenameWorker Fehler: {e}")
            self.error.emit(str(e))


class PDFSaveWorker(QThread):
    """Worker zum asynchronen Speichern eines bearbeiteten PDFs auf dem Server."""
    finished = Signal(bool)   # success
    error = Signal(str)       # error_message
    
    def __init__(self, docs_api, doc_id: int, file_path: str, parent=None):
        super().__init__(parent)
        self.docs_api = docs_api
        self.doc_id = doc_id
        self.file_path = file_path
    
    def run(self):
        try:
            success = self.docs_api.replace_document_file(self.doc_id, self.file_path)
            self.finished.emit(success)
        except Exception as e:
            self.error.emit(str(e))


class PDFRefreshWorker(QThread):
    """Aktualisiert Leere-Seiten-Daten und Textinhalt nach PDF-Bearbeitung."""
    finished = Signal(bool)
    
    def __init__(self, docs_api, doc_id: int, pdf_path: str, parent=None):
        super().__init__(parent)
        self.docs_api = docs_api
        self.doc_id = doc_id
        self.pdf_path = pdf_path
    
    def run(self):
        try:
            from services.empty_page_detector import get_empty_pages
            empty_indices, total_pages = get_empty_pages(self.pdf_path)
            if total_pages > 0:
                self.docs_api.client.put(
                    f'/documents/{self.doc_id}',
                    json_data={
                        'empty_page_count': len(empty_indices),
                        'total_page_count': total_pages
                    }
                )
            
            from services.early_text_extract import extract_and_save_text
            extract_and_save_text(self.docs_api, self.doc_id, self.pdf_path)
            
            self.finished.emit(True)
        except Exception as e:
            logger.error(f"PDF-Refresh nach Bearbeitung fehlgeschlagen fuer Dokument {self.doc_id}: {e}")
            self.finished.emit(False)


class PDFViewerDialog(QDialog):
    """
    Dialog zur PDF-Vorschau und -Bearbeitung.
    
    Zeigt PDFs direkt in der App an ohne separaten Download.
    Verwendet QPdfView (Qt6 native) oder öffnet extern als Fallback.
    
    Bearbeitungsmodus (editable=True):
    - Thumbnail-Sidebar links mit Seitenvorschau
    - Seiten drehen (CW/CCW) und loeschen
    - Bearbeitetes PDF auf dem Server speichern
    """
    
    # Signal wenn PDF gespeichert wurde (fuer Cache-Invalidierung)
    pdf_saved = Signal(int)  # doc_id
    
    def __init__(self, pdf_path: str, title: str = "PDF-Vorschau", parent=None,
                 doc_id: int = None, docs_api=None, editable: bool = False):
        super().__init__(parent)
        self.pdf_path = pdf_path
        self.pdf_document = None
        self._doc_id = doc_id
        self._docs_api = docs_api
        self._editable = editable and doc_id is not None and docs_api is not None
        self._fitz_doc = None
        self._change_count = 0
        self._save_worker = None
        self._refresh_data_worker = None
        self._temp_pdf_path = None
        self.setWindowTitle(title)
        self.setMinimumSize(900, 700)
        self.resize(1200 if self._editable else 1000, 900 if self._editable else 800)
        
        self._setup_ui()
        
        # PyMuPDF laden fuer Bearbeitung
        if self._editable:
            self._load_fitz_document()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        
        # Toolbar
        toolbar = QToolBar()
        toolbar.setMovable(False)
        
        # Titel
        title_label = QLabel(f"  {os.path.basename(self.pdf_path)}")
        title_label.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        toolbar.addWidget(title_label)
        
        toolbar.addSeparator()
        
        # Zoom-Buttons (für QPdfView)
        if HAS_PDF_VIEW:
            zoom_in_btn = QPushButton("Zoom +")
            zoom_in_btn.setToolTip("Vergroessern")
            zoom_in_btn.clicked.connect(self._zoom_in)
            toolbar.addWidget(zoom_in_btn)
            
            zoom_out_btn = QPushButton("Zoom -")
            zoom_out_btn.setToolTip("Verkleinern")
            zoom_out_btn.clicked.connect(self._zoom_out)
            toolbar.addWidget(zoom_out_btn)
            
            fit_width_btn = QPushButton("Breite")
            fit_width_btn.setToolTip("An Breite anpassen")
            fit_width_btn.clicked.connect(self._fit_width)
            toolbar.addWidget(fit_width_btn)
            
            fit_page_btn = QPushButton("Seite")
            fit_page_btn.setToolTip("Ganze Seite")
            fit_page_btn.clicked.connect(self._fit_page)
            toolbar.addWidget(fit_page_btn)
            
            toolbar.addSeparator()
        
        # Bearbeitungs-Buttons (nur im Edit-Modus)
        if self._editable:
            from i18n.de import (
                PDF_EDIT_ROTATE_CCW, PDF_EDIT_ROTATE_CW,
                PDF_EDIT_DELETE_PAGE, PDF_EDIT_SAVE
            )
            
            _edit_btn_style = """
                QPushButton {
                    padding: 4px 10px;
                    border: 1px solid #cbd5e1;
                    border-radius: 4px;
                    background-color: #f8fafc;
                    font-size: 11px;
                }
                QPushButton:hover { background-color: #e2e8f0; }
                QPushButton:pressed { background-color: #cbd5e1; }
            """
            _delete_btn_style = """
                QPushButton {
                    padding: 4px 10px;
                    border: 1px solid #fca5a5;
                    border-radius: 4px;
                    background-color: #fef2f2;
                    color: #dc2626;
                    font-size: 11px;
                }
                QPushButton:hover { background-color: #fee2e2; }
                QPushButton:pressed { background-color: #fecaca; }
            """
            
            rotate_ccw_btn = QPushButton(PDF_EDIT_ROTATE_CCW)
            rotate_ccw_btn.setToolTip(PDF_EDIT_ROTATE_CCW)
            rotate_ccw_btn.setStyleSheet(_edit_btn_style)
            rotate_ccw_btn.clicked.connect(self._rotate_ccw)
            toolbar.addWidget(rotate_ccw_btn)
            
            rotate_cw_btn = QPushButton(PDF_EDIT_ROTATE_CW)
            rotate_cw_btn.setToolTip(PDF_EDIT_ROTATE_CW)
            rotate_cw_btn.setStyleSheet(_edit_btn_style)
            rotate_cw_btn.clicked.connect(self._rotate_cw)
            toolbar.addWidget(rotate_cw_btn)
            
            self._delete_page_btn = QPushButton(PDF_EDIT_DELETE_PAGE)
            self._delete_page_btn.setToolTip(PDF_EDIT_DELETE_PAGE)
            self._delete_page_btn.setStyleSheet(_delete_btn_style)
            self._delete_page_btn.clicked.connect(self._delete_page)
            toolbar.addWidget(self._delete_page_btn)
            
            toolbar.addSeparator()
            
            self._save_btn = QPushButton(PDF_EDIT_SAVE)
            self._save_btn.setToolTip(PDF_EDIT_SAVE)
            self._save_btn.setEnabled(False)
            self._save_btn.setStyleSheet("""
                QPushButton {
                    background-color: #059669;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 4px 12px;
                    font-weight: bold;
                }
                QPushButton:hover { background-color: #047857; }
                QPushButton:disabled { background-color: #9ca3af; }
            """)
            self._save_btn.clicked.connect(self._save_pdf)
            toolbar.addWidget(self._save_btn)
            
            toolbar.addSeparator()
        
        # Spacer um die rechten Elemente nach rechts zu druecken
        if self._editable:
            from PySide6.QtWidgets import QSizePolicy
            spacer = QWidget()
            spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
            toolbar.addWidget(spacer)
            
            from i18n.de import PDF_EDIT_NO_CHANGES
            self._edit_status_label = QLabel(PDF_EDIT_NO_CHANGES)
            self._edit_status_label.setStyleSheet(
                "color: #6b7280; font-size: 11px; padding: 0 8px;"
            )
            toolbar.addWidget(self._edit_status_label)
            
            toolbar.addSeparator()
        
        # Extern oeffnen
        open_external_btn = QPushButton("Extern oeffnen")
        open_external_btn.setToolTip("Mit System-PDF-Viewer oeffnen")
        open_external_btn.clicked.connect(self._open_external)
        toolbar.addWidget(open_external_btn)
        
        # Schließen
        close_btn = QPushButton("Schliessen")
        close_btn.clicked.connect(self.close)
        toolbar.addWidget(close_btn)
        
        layout.addWidget(toolbar)
        
        # Inline-Status-Label fuer Fehler (statt modaler Dialoge)
        self._status_label = QLabel("")
        self._status_label.setVisible(False)
        self._status_label.setWordWrap(True)
        layout.addWidget(self._status_label)
        
        # Hauptbereich: Thumbnails (optional) + PDF-Viewer
        if HAS_PDF_VIEW:
            if self._editable:
                # Splitter: Thumbnails links, QPdfView rechts
                from PySide6.QtWidgets import QSplitter, QListWidget, QListWidgetItem
                splitter = QSplitter(Qt.Orientation.Horizontal)
                
                # Thumbnail-Liste mit Mehrfachauswahl (Strg+Klick, Shift+Klick, Strg+A)
                self._thumbnail_list = QListWidget()
                from PySide6.QtWidgets import QAbstractItemView
                from PySide6.QtCore import QSize
                self._thumbnail_list.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
                self._thumbnail_list.setIconSize(QSize(120, 160))
                self._thumbnail_list.setMinimumWidth(145)
                self._thumbnail_list.setMaximumWidth(160)
                self._thumbnail_list.setSpacing(2)
                self._thumbnail_list.setStyleSheet("""
                    QListWidget {
                        background-color: #f1f5f9;
                        border: none;
                        border-right: 1px solid #e2e8f0;
                        font-size: 10px;
                    }
                    QListWidget::item {
                        padding: 3px;
                        border-radius: 3px;
                    }
                    QListWidget::item:selected {
                        background-color: #dbeafe;
                        border: 2px solid #3b82f6;
                    }
                """)
                self._thumbnail_list.itemSelectionChanged.connect(self._on_thumbnail_selection_changed)
                splitter.addWidget(self._thumbnail_list)
                
                # QPdfView
                self.pdf_document = QPdfDocument(self)
                self.pdf_view = QPdfView(self)
                self.pdf_view.setDocument(self.pdf_document)
                self.pdf_view.setPageMode(QPdfView.PageMode.MultiPage)
                self.pdf_view.setZoomMode(QPdfView.ZoomMode.FitInView)
                splitter.addWidget(self.pdf_view)
                
                splitter.setSizes([150, 950])
                
                layout.addWidget(splitter)
            else:
                # Read-only: Nur QPdfView (wie bisher)
                self.pdf_document = QPdfDocument(self)
                self.pdf_view = QPdfView(self)
                self.pdf_view.setDocument(self.pdf_document)
                self.pdf_view.setPageMode(QPdfView.PageMode.MultiPage)
                self.pdf_view.setZoomMode(QPdfView.ZoomMode.FitInView)
                layout.addWidget(self.pdf_view)
            
            # PDF laden
            error = self.pdf_document.load(self.pdf_path)
            if error != QPdfDocument.Error.None_:
                self._status_label.setText(f"PDF konnte nicht geladen werden: {error}")
                self._status_label.setStyleSheet(
                    "color: #dc2626; background: #fef2f2; padding: 6px 12px; border-radius: 4px;"
                )
                self._status_label.setVisible(True)
            
            self._zoom_factor = 1.0
        else:
            # Fallback: Hinweis und Button zum externen Öffnen
            fallback_widget = QWidget()
            fallback_layout = QVBoxLayout(fallback_widget)
            fallback_layout.addStretch()
            
            info_label = QLabel(
                "PDF-Vorschau nicht verfuegbar.\n\n"
                "Fuer die integrierte PDF-Ansicht wird\n"
                "PySide6 >= 6.4 benoetigt.\n\n"
                "Das PDF wird extern geoeffnet."
            )
            info_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            info_label.setFont(QFont("Segoe UI", 11))
            fallback_layout.addWidget(info_label)
            
            open_btn = QPushButton("PDF extern oeffnen")
            open_btn.setMinimumSize(200, 50)
            open_btn.clicked.connect(self._open_external)
            fallback_layout.addWidget(open_btn, alignment=Qt.AlignmentFlag.AlignCenter)
            
            # Automatisch extern öffnen wenn kein Viewer verfügbar
            self._open_external()
            
            fallback_layout.addStretch()
            layout.addWidget(fallback_widget)
    
    # ========================================
    # PyMuPDF Integration (Bearbeitungsmodus)
    # ========================================
    
    def _load_fitz_document(self):
        """Laedt das PDF mit PyMuPDF fuer Bearbeitungsoperationen."""
        try:
            import fitz
            try:
                self._fitz_doc = fitz.open(self.pdf_path)
            except Exception:
                # Fallback: Datei als Bytes laden (Workaround fuer MuPDF-Probleme
                # mit Sonderzeichen wie '...' in Windows-Pfaden)
                logger.warning(f"PyMuPDF Pfad-Oeffnung fehlgeschlagen, versuche Bytes-Fallback: {self.pdf_path}")
                with open(self.pdf_path, 'rb') as f:
                    data = f.read()
                self._fitz_doc = fitz.open(stream=data, filetype="pdf")
            self._refresh_thumbnails()
        except Exception as e:
            logger.error(f"PyMuPDF konnte PDF nicht laden: {e}")
            self._status_label.setText(f"PDF-Bearbeitung nicht verfuegbar: {e}")
            self._status_label.setStyleSheet(
                "color: #dc2626; background: #fef2f2; padding: 6px 12px; border-radius: 4px;"
            )
            self._status_label.setVisible(True)
    
    def _refresh_thumbnails(self):
        """Rendert alle Seiten als Thumbnails in die Sidebar.
        
        Signale werden waehrend des Rebuilds blockiert, damit der Caller
        die Auswahl selbst steuern kann (Einzel- oder Mehrfachauswahl).
        """
        if not self._fitz_doc or not hasattr(self, '_thumbnail_list'):
            return
        
        from PySide6.QtGui import QPixmap, QImage, QIcon
        from PySide6.QtCore import QSize
        from PySide6.QtWidgets import QListWidgetItem
        
        self._thumbnail_list.blockSignals(True)
        self._thumbnail_list.clear()
        
        for i in range(len(self._fitz_doc)):
            page = self._fitz_doc[i]
            # Thumbnail rendern (120px Breite)
            zoom = 120.0 / page.rect.width
            mat = __import__('fitz').Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            
            # PyMuPDF Pixmap -> QImage -> QPixmap -> QIcon
            img = QImage(pix.samples, pix.width, pix.height, pix.stride, QImage.Format.Format_RGB888)
            pixmap = QPixmap.fromImage(img)
            
            item = QListWidgetItem()
            item.setIcon(QIcon(pixmap))
            item.setText(f"S. {i + 1}")
            item.setData(Qt.ItemDataRole.UserRole, i)  # Seiten-Index
            self._thumbnail_list.addItem(item)
        
        self._thumbnail_list.setIconSize(QSize(120, 160))
        self._thumbnail_list.blockSignals(False)
        
        # Erste Seite auswaehlen (wenn kein externer Caller die Auswahl uebernimmt)
        if self._thumbnail_list.count() > 0:
            self._thumbnail_list.setCurrentRow(0)
    
    def _get_selected_page_indices(self) -> list:
        """Gibt die Indizes aller ausgewaehlten Seiten zurueck (sortiert)."""
        if not hasattr(self, '_thumbnail_list'):
            return []
        indices = []
        for item in self._thumbnail_list.selectedItems():
            idx = item.data(Qt.ItemDataRole.UserRole)
            if idx is not None:
                indices.append(idx)
        return sorted(indices)
    
    def _on_thumbnail_selection_changed(self):
        """Scrollt die PDF-Ansicht zur zuletzt ausgewaehlten Seite."""
        if not hasattr(self, 'pdf_view'):
            return
        indices = self._get_selected_page_indices()
        if not indices:
            return
        try:
            from PySide6.QtCore import QPointF
            navigator = self.pdf_view.pageNavigator()
            navigator.jump(indices[-1], QPointF(0, 0))
        except Exception as e:
            logger.debug(f"Seiten-Navigation fehlgeschlagen: {e}")
        self._update_edit_status()
    
    def _rotate_cw(self):
        """Dreht die ausgewaehlte Seite 90 Grad im Uhrzeigersinn."""
        self._rotate_page(90)
    
    def _rotate_ccw(self):
        """Dreht die ausgewaehlte Seite 90 Grad gegen den Uhrzeigersinn."""
        self._rotate_page(-90)
    
    def _rotate_page(self, degrees: int):
        """Dreht alle ausgewaehlten Seiten um die angegebenen Grad."""
        if not self._fitz_doc:
            return
        
        page_indices = self._get_selected_page_indices()
        if not page_indices:
            return
        
        for idx in page_indices:
            page = self._fitz_doc[idx]
            page.set_rotation((page.rotation + degrees) % 360)
        
        self._change_count += len(page_indices)
        self._apply_changes_and_refresh(page_indices[0], selected_pages=page_indices)
    
    def _delete_page(self):
        """Loescht alle ausgewaehlten Seiten nach Bestaetigung."""
        if not self._fitz_doc:
            return
        
        from i18n.de import (PDF_EDIT_DELETE_CONFIRM, PDF_EDIT_DELETE_MULTI_CONFIRM,
                              PDF_EDIT_MIN_ONE_PAGE)
        
        page_indices = self._get_selected_page_indices()
        if not page_indices:
            return
        
        # Pruefung: Mindestens eine Seite muss verbleiben
        if len(page_indices) >= len(self._fitz_doc):
            self._status_label.setText(PDF_EDIT_MIN_ONE_PAGE)
            self._status_label.setStyleSheet(
                "color: #f59e0b; background: #fffbeb; padding: 6px 12px; border-radius: 4px;"
            )
            self._status_label.setVisible(True)
            return
        
        # Bestaetigung (Einzel- vs. Mehrfachauswahl)
        if len(page_indices) == 1:
            confirm_msg = PDF_EDIT_DELETE_CONFIRM.format(page=page_indices[0] + 1)
        else:
            confirm_msg = PDF_EDIT_DELETE_MULTI_CONFIRM.format(count=len(page_indices))
        
        reply = QMessageBox.question(
            self, "Seite loeschen",
            confirm_msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # Loeschen in umgekehrter Reihenfolge (hoechster Index zuerst),
        # damit sich die Indizes der noch zu loeschenden Seiten nicht verschieben.
        for idx in sorted(page_indices, reverse=True):
            self._fitz_doc.delete_page(idx)
        
        self._change_count += len(page_indices)
        
        # Neue Auswahl: erste verbleibende Seite nach der ersten geloeschten
        new_idx = min(page_indices[0], len(self._fitz_doc) - 1)
        self._apply_changes_and_refresh(new_idx)
    
    def _apply_changes_and_refresh(self, select_page: int = 0,
                                    selected_pages: list = None):
        """Speichert das geaenderte PDF temporaer und aktualisiert die Anzeige.
        
        Args:
            select_page: Einzelne Seite die ausgewaehlt werden soll (Fallback)
            selected_pages: Liste von Seiten die nach Refresh selektiert werden sollen
        """
        if not self._fitz_doc:
            return
        
        import fitz
        
        # Alternierende Temp-Dateien: A und B
        # PyMuPDF kann nicht an die gleiche Datei speichern, von der es gelesen hat.
        temp_dir = tempfile.gettempdir()
        old_path = self._temp_pdf_path
        
        # Zwischen A und B wechseln
        path_a = os.path.join(temp_dir, f'bipro_edit_{os.getpid()}_a.pdf')
        path_b = os.path.join(temp_dir, f'bipro_edit_{os.getpid()}_b.pdf')
        
        if old_path == path_a:
            new_path = path_b
        else:
            new_path = path_a
        
        self._fitz_doc.save(new_path)
        self._temp_pdf_path = new_path
        
        # QPdfView neu laden
        if HAS_PDF_VIEW and self.pdf_document:
            self.pdf_document.close()
            self.pdf_document.load(new_path)
        
        # fitz-Dokument neu laden
        self._fitz_doc.close()
        self._fitz_doc = fitz.open(new_path)
        
        # Alte Temp-Datei aufraemen
        if old_path and old_path != new_path and os.path.exists(old_path):
            try:
                os.unlink(old_path)
            except Exception:
                pass
        
        # Thumbnails aktualisieren
        self._refresh_thumbnails()
        
        # Seiten-Auswahl wiederherstellen
        if hasattr(self, '_thumbnail_list'):
            page_count = self._thumbnail_list.count()
            if selected_pages and len(selected_pages) > 1:
                # Mehrfachauswahl wiederherstellen
                self._thumbnail_list.blockSignals(True)
                self._thumbnail_list.clearSelection()
                for idx in selected_pages:
                    if idx < page_count:
                        self._thumbnail_list.item(idx).setSelected(True)
                self._thumbnail_list.blockSignals(False)
                self._on_thumbnail_selection_changed()
            elif select_page < page_count:
                self._thumbnail_list.setCurrentRow(select_page)
        
        # UI aktualisieren
        self._update_edit_status()
        self._status_label.setVisible(False)
    
    def _update_edit_status(self):
        """Aktualisiert die Statusbar mit Aenderungszaehler."""
        if not hasattr(self, '_edit_status_label'):
            return
        
        from i18n.de import (PDF_EDIT_CHANGES, PDF_EDIT_NO_CHANGES, PDF_EDIT_STATUS,
                              PDF_EDIT_MULTI_SELECTED)
        
        page_count = len(self._fitz_doc) if self._fitz_doc else 0
        indices = self._get_selected_page_indices()
        
        parts = []
        if page_count > 0 and indices:
            if len(indices) == 1:
                parts.append(PDF_EDIT_STATUS.format(current=indices[0] + 1, total=page_count))
            else:
                parts.append(PDF_EDIT_MULTI_SELECTED.format(
                    selected=len(indices), total=page_count))
        
        if self._change_count > 0:
            parts.append(PDF_EDIT_CHANGES.format(count=self._change_count))
        else:
            parts.append(PDF_EDIT_NO_CHANGES)
        
        self._edit_status_label.setText("  |  ".join(parts))
        
        # Save-Button aktivieren wenn Aenderungen vorhanden
        if hasattr(self, '_save_btn'):
            self._save_btn.setEnabled(self._change_count > 0)
    
    def _save_pdf(self):
        """Speichert das bearbeitete PDF auf dem Server."""
        if not self._fitz_doc or not self._docs_api or self._doc_id is None:
            return
        
        from i18n.de import PDF_EDIT_SAVING
        
        # Finales PDF in temp-Datei speichern
        save_path = os.path.join(tempfile.gettempdir(), f'bipro_save_{self._doc_id}.pdf')
        self._fitz_doc.save(save_path, garbage=4, deflate=True)
        
        # Save-Button deaktivieren waehrend Upload
        self._save_btn.setEnabled(False)
        self._save_btn.setText(PDF_EDIT_SAVING)
        
        # Worker starten
        self._save_worker = PDFSaveWorker(self._docs_api, self._doc_id, save_path)
        self._save_worker.finished.connect(self._on_save_finished)
        self._save_worker.error.connect(self._on_save_error)
        self._save_worker.start()
    
    def _on_save_finished(self, success: bool):
        """Callback nach dem Speichern."""
        from i18n.de import (PDF_EDIT_SAVE, PDF_EDIT_SAVE_SUCCESS, PDF_EDIT_SAVE_ERROR,
                              PDF_EDIT_REFRESHING)
        
        self._save_btn.setText(PDF_EDIT_SAVE)
        
        if success:
            self._change_count = 0
            self._update_edit_status()
            self._status_label.setText(PDF_EDIT_SAVE_SUCCESS)
            self._status_label.setStyleSheet(
                "color: #059669; background: #ecfdf5; padding: 6px 12px; border-radius: 4px;"
            )
            self._status_label.setVisible(True)
            # Signal fuer Cache-Invalidierung
            self.pdf_saved.emit(self._doc_id)
            
            # Leere-Seiten + Textinhalt im Hintergrund aktualisieren
            if self._docs_api and self._doc_id is not None and self._temp_pdf_path:
                self._status_label.setText(PDF_EDIT_REFRESHING)
                self._status_label.setStyleSheet(
                    "color: #2563eb; background: #eff6ff; padding: 6px 12px; border-radius: 4px;"
                )
                self._refresh_data_worker = PDFRefreshWorker(
                    self._docs_api, self._doc_id, self._temp_pdf_path
                )
                self._refresh_data_worker.finished.connect(self._on_refresh_data_finished)
                self._refresh_data_worker.start()
        else:
            self._save_btn.setEnabled(True)
            self._status_label.setText(PDF_EDIT_SAVE_ERROR.format(error="Server-Fehler"))
            self._status_label.setStyleSheet(
                "color: #dc2626; background: #fef2f2; padding: 6px 12px; border-radius: 4px;"
            )
            self._status_label.setVisible(True)
    
    def _on_refresh_data_finished(self, success: bool):
        """Callback nach dem Aktualisieren von Leere-Seiten und Textinhalt."""
        from i18n.de import PDF_EDIT_REFRESH_SUCCESS, PDF_EDIT_SAVE_SUCCESS
        
        if success:
            self._status_label.setText(PDF_EDIT_REFRESH_SUCCESS)
            self._status_label.setStyleSheet(
                "color: #059669; background: #ecfdf5; padding: 6px 12px; border-radius: 4px;"
            )
        else:
            self._status_label.setText(PDF_EDIT_SAVE_SUCCESS)
            self._status_label.setStyleSheet(
                "color: #059669; background: #ecfdf5; padding: 6px 12px; border-radius: 4px;"
            )
    
    def _on_save_error(self, error_msg: str):
        """Callback bei Speicher-Fehler."""
        from i18n.de import PDF_EDIT_SAVE, PDF_EDIT_SAVE_ERROR
        
        self._save_btn.setText(PDF_EDIT_SAVE)
        self._save_btn.setEnabled(True)
        self._status_label.setText(PDF_EDIT_SAVE_ERROR.format(error=error_msg))
        self._status_label.setStyleSheet(
            "color: #dc2626; background: #fef2f2; padding: 6px 12px; border-radius: 4px;"
        )
        self._status_label.setVisible(True)
    
    # ========================================
    # Zoom (QPdfView)
    # ========================================
    
    def _zoom_in(self):
        if HAS_PDF_VIEW:
            self._zoom_factor = min(4.0, self._zoom_factor + 0.25)
            self.pdf_view.setZoomFactor(self._zoom_factor)
            self.pdf_view.setZoomMode(QPdfView.ZoomMode.Custom)
    
    def _zoom_out(self):
        if HAS_PDF_VIEW:
            self._zoom_factor = max(0.25, self._zoom_factor - 0.25)
            self.pdf_view.setZoomFactor(self._zoom_factor)
            self.pdf_view.setZoomMode(QPdfView.ZoomMode.Custom)
    
    def _fit_width(self):
        if HAS_PDF_VIEW:
            self.pdf_view.setZoomMode(QPdfView.ZoomMode.FitToWidth)
            self._zoom_factor = self.pdf_view.zoomFactor()
    
    def _fit_page(self):
        if HAS_PDF_VIEW:
            self.pdf_view.setZoomMode(QPdfView.ZoomMode.FitInView)
            self._zoom_factor = self.pdf_view.zoomFactor()
    
    def _open_external(self):
        """Oeffnet das PDF mit dem System-Standard-Viewer."""
        import subprocess
        import sys
        
        try:
            if sys.platform == 'win32':
                os.startfile(self.pdf_path)
            elif sys.platform == 'darwin':
                subprocess.run(['open', self.pdf_path])
            else:
                subprocess.run(['xdg-open', self.pdf_path])
        except Exception as e:
            self._status_label.setText(f"Konnte PDF nicht oeffnen: {e}")
            self._status_label.setStyleSheet(
                "color: #dc2626; background: #fef2f2; padding: 6px 12px; border-radius: 4px;"
            )
            self._status_label.setVisible(True)
    
    def showEvent(self, event):
        """Maximiert den Dialog im Bearbeitungsmodus beim ersten Anzeigen."""
        super().showEvent(event)
        if self._editable and not getattr(self, '_was_shown', False):
            self._was_shown = True
            self.showMaximized()
    
    def closeEvent(self, event):
        """Warnt bei ungespeicherten Aenderungen."""
        if self._change_count > 0:
            from i18n.de import PDF_EDIT_UNSAVED, PDF_EDIT_UNSAVED_CONFIRM
            reply = QMessageBox.question(
                self, PDF_EDIT_UNSAVED, PDF_EDIT_UNSAVED_CONFIRM,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                event.ignore()
                return
        
        # Laufende Worker sauber beenden bevor der Dialog zerstoert wird
        for worker_attr in ('_save_worker', '_refresh_data_worker'):
            worker = getattr(self, worker_attr, None)
            if worker and worker.isRunning():
                worker.quit()
                worker.wait(3000)
        
        # Cleanup
        if self._fitz_doc:
            try:
                self._fitz_doc.close()
            except Exception:
                pass
        
        # Temp-Dateien aufraemen (beide alternierenden + save-Datei)
        temp_dir = tempfile.gettempdir()
        for suffix in ['_a.pdf', '_b.pdf']:
            p = os.path.join(temp_dir, f'bipro_edit_{os.getpid()}{suffix}')
            if os.path.exists(p):
                try:
                    os.unlink(p)
                except Exception:
                    pass
        
        save_path = os.path.join(temp_dir, f'bipro_save_{self._doc_id}.pdf') if self._doc_id else None
        if save_path and os.path.exists(save_path):
            try:
                os.unlink(save_path)
            except Exception:
                pass
        
        super().closeEvent(event)


# ========================================
# Excel/CSV-Vorschau
# ========================================

# openpyxl fuer .xlsx (optional)
HAS_OPENPYXL = False
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    pass


class SpreadsheetViewerDialog(QDialog):
    """
    Dialog zur Vorschau von CSV- und Excel-Dateien.
    
    Zeigt tabellarische Daten in einem QTableWidget an.
    Unterstuetzt:
    - CSV (.csv) via Python csv-Modul
    - Excel (.xlsx) via openpyxl
    - TSV (.tsv) via Python csv-Modul
    """
    
    # Maximale Zeilen fuer die Vorschau (Performance-Schutz)
    MAX_PREVIEW_ROWS = 5000
    
    def __init__(self, file_path: str, title: str = "", parent=None):
        super().__init__(parent)
        self.file_path = file_path
        self._sheets_data: dict = {}  # {sheet_name: (headers, rows)}
        self._current_sheet: str = ""
        
        from i18n.de import SPREADSHEET_PREVIEW_TITLE
        display_title = title or SPREADSHEET_PREVIEW_TITLE.format(
            filename=os.path.basename(file_path)
        )
        self.setWindowTitle(display_title)
        self.setMinimumSize(900, 600)
        self.resize(1100, 750)
        
        self._setup_ui()
        self._load_data()
    
    def _setup_ui(self):
        """Erstellt die UI-Elemente."""
        from i18n.de import (
            SPREADSHEET_SHEET_LABEL, SPREADSHEET_EXTERN_OPEN,
            SPREADSHEET_CLOSE, SPREADSHEET_NO_DATA
        )
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        
        # Toolbar
        toolbar = QToolBar()
        toolbar.setMovable(False)
        
        # Titel
        title_label = QLabel(f"  {os.path.basename(self.file_path)}")
        title_label.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        toolbar.addWidget(title_label)
        
        toolbar.addSeparator()
        
        # Sheet-Auswahl (nur fuer Excel mit mehreren Blaettern)
        self._sheet_label = QLabel(f"  {SPREADSHEET_SHEET_LABEL} ")
        self._sheet_label.setVisible(False)
        toolbar.addWidget(self._sheet_label)
        
        self._sheet_combo = QComboBox()
        self._sheet_combo.setMinimumWidth(150)
        self._sheet_combo.currentTextChanged.connect(self._on_sheet_changed)
        self._sheet_combo.setVisible(False)
        toolbar.addWidget(self._sheet_combo)
        
        toolbar.addSeparator()
        
        # Zeilen-Info
        self._info_label = QLabel("")
        self._info_label.setFont(QFont("Segoe UI", 9))
        toolbar.addWidget(self._info_label)
        
        # Spacer
        spacer = QWidget()
        from PySide6.QtWidgets import QSizePolicy
        spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        toolbar.addWidget(spacer)
        
        # Extern oeffnen
        open_external_btn = QPushButton(SPREADSHEET_EXTERN_OPEN)
        open_external_btn.setToolTip("Mit System-Anwendung oeffnen")
        open_external_btn.clicked.connect(self._open_external)
        toolbar.addWidget(open_external_btn)
        
        toolbar.addSeparator()
        
        # Schliessen
        close_btn = QPushButton(SPREADSHEET_CLOSE)
        close_btn.clicked.connect(self.close)
        toolbar.addWidget(close_btn)
        
        layout.addWidget(toolbar)
        
        # Tabelle
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setDefaultSectionSize(24)
        self.table.setStyleSheet("""
            QTableWidget {
                font-family: 'Segoe UI', 'Consolas', monospace;
                font-size: 12px;
                gridline-color: #e0e0e0;
            }
            QTableWidget::item {
                padding: 2px 6px;
            }
            QHeaderView::section {
                background-color: #f0f0f0;
                padding: 4px 6px;
                border: 1px solid #d0d0d0;
                font-weight: bold;
                font-size: 11px;
            }
            QTableWidget::item:alternate {
                background-color: #fafafa;
            }
        """)
        layout.addWidget(self.table)
        
        # Hinweis-Label (bei leeren Daten)
        self._empty_label = QLabel(SPREADSHEET_NO_DATA)
        self._empty_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._empty_label.setFont(QFont("Segoe UI", 11))
        self._empty_label.setVisible(False)
        layout.addWidget(self._empty_label)
        
        # Inline-Status-Label fuer Fehler/Hinweise (statt modaler Dialoge)
        self._status_label = QLabel("")
        self._status_label.setVisible(False)
        self._status_label.setWordWrap(True)
        layout.addWidget(self._status_label)
    
    def _load_data(self):
        """Laedt die Daten basierend auf Dateiendung."""
        ext = os.path.splitext(self.file_path)[1].lower()
        
        try:
            if ext == '.csv':
                self._load_csv()
            elif ext == '.tsv':
                self._load_csv(delimiter='\t')
            elif ext == '.xlsx':
                self._load_xlsx()
            elif ext == '.xls':
                self._show_xls_message()
                return
            else:
                # Versuche als CSV (z.B. fuer .txt mit Tabulator-Trennung)
                self._load_csv()
        except Exception as e:
            from i18n.de import SPREADSHEET_LOAD_ERROR
            logger.error(f"Fehler beim Laden der Tabelle: {e}")
            self._status_label.setText(SPREADSHEET_LOAD_ERROR.format(error=str(e)))
            self._status_label.setStyleSheet(
                "color: #dc2626; background: #fef2f2; padding: 6px 12px; border-radius: 4px;"
            )
            self._status_label.setVisible(True)
            return
        
        # Erstes Sheet anzeigen
        if self._sheets_data:
            sheet_names = list(self._sheets_data.keys())
            
            # Sheet-Auswahl nur bei mehreren Blaettern anzeigen
            if len(sheet_names) > 1:
                self._sheet_label.setVisible(True)
                self._sheet_combo.setVisible(True)
                self._sheet_combo.addItems(sheet_names)
            
            self._display_sheet(sheet_names[0])
    
    def _load_csv(self, delimiter: str = None):
        """
        Laedt CSV-Datei mit automatischer Delimiter- und Encoding-Erkennung.
        
        Args:
            delimiter: Optionaler Delimiter. Wenn None, wird automatisch erkannt.
        """
        import csv
        
        # Encoding-Reihenfolge (wie beim GDV-Parser)
        encodings = ['utf-8-sig', 'utf-8', 'cp1252', 'latin-1']
        content = None
        used_encoding = None
        
        for enc in encodings:
            try:
                with open(self.file_path, 'r', encoding=enc) as f:
                    content = f.read()
                    used_encoding = enc
                    break
            except (UnicodeDecodeError, UnicodeError):
                continue
        
        if content is None:
            from i18n.de import SPREADSHEET_ENCODING_ERROR
            raise ValueError(SPREADSHEET_ENCODING_ERROR)
        
        # Delimiter automatisch erkennen wenn nicht vorgegeben
        if delimiter is None:
            sniffer = csv.Sniffer()
            try:
                # Erste 8KB fuer Erkennung verwenden
                sample = content[:8192]
                dialect = sniffer.sniff(sample, delimiters=',;\t|')
                delimiter = dialect.delimiter
            except csv.Error:
                # Fallback: Semikolon (deutsch) oder Komma
                delimiter = ';' if ';' in content[:2000] else ','
        
        # CSV parsen
        lines = content.splitlines()
        reader = csv.reader(lines, delimiter=delimiter)
        
        rows = []
        headers = []
        
        for i, row in enumerate(reader):
            if i == 0:
                # Erste Zeile als Header verwenden
                headers = [str(h).strip() for h in row]
            else:
                rows.append([str(cell) for cell in row])
            
            if i >= self.MAX_PREVIEW_ROWS:
                break
        
        # Falls keine Header erkannt (z.B. nur Daten)
        if not headers and rows:
            # Generische Header
            max_cols = max(len(r) for r in rows) if rows else 0
            headers = [f"Spalte {i+1}" for i in range(max_cols)]
        
        sheet_name = os.path.basename(self.file_path)
        self._sheets_data[sheet_name] = (headers, rows, len(lines) - 1)
    
    def _load_xlsx(self):
        """Laedt Excel-Datei (.xlsx) via openpyxl."""
        if not HAS_OPENPYXL:
            from i18n.de import SPREADSHEET_XLSX_NOT_AVAILABLE
            self._status_label.setText(SPREADSHEET_XLSX_NOT_AVAILABLE)
            self._status_label.setStyleSheet(
                "color: #1e40af; background: #eff6ff; padding: 6px 12px; border-radius: 4px;"
            )
            self._status_label.setVisible(True)
            self._open_external()
            return
        
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            rows = []
            total_rows = 0
            
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    # Erste Zeile als Header
                    headers = [str(cell) if cell is not None else "" for cell in row]
                else:
                    rows.append([
                        str(cell) if cell is not None else "" for cell in row
                    ])
                
                total_rows = i
                
                if i >= self.MAX_PREVIEW_ROWS:
                    break
            
            # Falls Sheet leer ist
            if not headers and not rows:
                headers = ["(leer)"]
            
            # Falls keine Header erkannt
            if not headers and rows:
                max_cols = max(len(r) for r in rows) if rows else 0
                headers = [f"Spalte {i+1}" for i in range(max_cols)]
            
            self._sheets_data[sheet_name] = (headers, rows, total_rows)
        
        wb.close()
    
    def _show_xls_message(self):
        """Zeigt Hinweis fuer alte .xls Dateien."""
        from i18n.de import SPREADSHEET_XLS_NOT_SUPPORTED
        self._status_label.setText(SPREADSHEET_XLS_NOT_SUPPORTED)
        self._status_label.setStyleSheet(
            "color: #1e40af; background: #eff6ff; padding: 6px 12px; border-radius: 4px;"
        )
        self._status_label.setVisible(True)
        self._open_external()
    
    def _display_sheet(self, sheet_name: str):
        """Zeigt die Daten eines Sheets in der Tabelle an."""
        from i18n.de import (
            SPREADSHEET_ROWS_INFO, SPREADSHEET_MAX_ROWS_INFO,
            SPREADSHEET_NO_DATA
        )
        
        if sheet_name not in self._sheets_data:
            return
        
        self._current_sheet = sheet_name
        headers, rows, total_rows = self._sheets_data[sheet_name]
        
        if not headers and not rows:
            self.table.setVisible(False)
            self._empty_label.setVisible(True)
            self._info_label.setText(SPREADSHEET_NO_DATA)
            return
        
        self.table.setVisible(True)
        self._empty_label.setVisible(False)
        
        # Tabelle befuellen
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(rows))
        
        for row_idx, row_data in enumerate(rows):
            for col_idx, cell_value in enumerate(row_data):
                if col_idx < len(headers):
                    item = QTableWidgetItem(cell_value)
                    self.table.setItem(row_idx, col_idx, item)
        
        # Spaltenbreiten anpassen
        self.table.resizeColumnsToContents()
        
        # Maximale Spaltenbreite begrenzen
        for col in range(self.table.columnCount()):
            if self.table.columnWidth(col) > 300:
                self.table.setColumnWidth(col, 300)
        
        # Info-Label aktualisieren
        shown_rows = len(rows)
        if total_rows > self.MAX_PREVIEW_ROWS:
            self._info_label.setText(
                f"  {SPREADSHEET_MAX_ROWS_INFO.format(shown=shown_rows, total=total_rows)}"
            )
        else:
            self._info_label.setText(
                f"  {SPREADSHEET_ROWS_INFO.format(rows=shown_rows, cols=len(headers))}"
            )
    
    def _on_sheet_changed(self, sheet_name: str):
        """Handler fuer Sheet-Wechsel."""
        if sheet_name and sheet_name in self._sheets_data:
            self._display_sheet(sheet_name)
    
    def _open_external(self):
        """Oeffnet die Datei mit der System-Anwendung."""
        import subprocess
        import sys
        
        try:
            if sys.platform == 'win32':
                os.startfile(self.file_path)
            elif sys.platform == 'darwin':
                subprocess.run(['open', self.file_path])
            else:
                subprocess.run(['xdg-open', self.file_path])
        except Exception as e:
            self._status_label.setText(f"Konnte Datei nicht oeffnen: {e}")
            self._status_label.setStyleSheet(
                "color: #dc2626; background: #fef2f2; padding: 6px 12px; border-radius: 4px;"
            )
            self._status_label.setVisible(True)


class ArchiveView(QWidget):
    """
    Dokumentenarchiv-Ansicht.
    
    Zeigt alle Dokumente vom Server mit Filter- und Such-Funktionen.
    """
    
    # Signal wenn ein GDV-Dokument geöffnet werden soll
    open_gdv_requested = Signal(int, str)  # doc_id, original_filename
    
    def __init__(self, api_client: APIClient, parent=None):
        super().__init__(parent)
        
        self.api_client = api_client
        self.docs_api = DocumentsAPI(api_client)
        
        self._documents: List[Document] = []
        self._load_worker = None
        self._upload_worker = None
        self._ai_rename_worker = None
        
        self._setup_ui()
        self.refresh_documents()
    
    def _setup_ui(self):
        """UI aufbauen."""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)
        
        # Header
        header_layout = QHBoxLayout()
        
        title = QLabel("Dokumentenarchiv")
        title.setFont(QFont("Segoe UI", 16, QFont.Weight.Bold))
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        # Refresh-Button
        refresh_btn = QPushButton("🔄 Aktualisieren")
        refresh_btn.clicked.connect(self.refresh_documents)
        header_layout.addWidget(refresh_btn)
        
        # Vorschau-Button
        self.preview_btn = QPushButton("👁️ Vorschau")
        self.preview_btn.setToolTip("PDF-Vorschau (Doppelklick auf PDF)")
        self.preview_btn.clicked.connect(self._preview_selected)
        header_layout.addWidget(self.preview_btn)
        
        # Download-Button fuer Mehrfachauswahl
        self.download_selected_btn = QPushButton("Ausgewaehlte herunterladen")
        self.download_selected_btn.clicked.connect(self._download_selected)
        header_layout.addWidget(self.download_selected_btn)
        
        # KI-Benennung Button
        self.ai_rename_btn = QPushButton("KI-Benennung")
        self.ai_rename_btn.setToolTip(
            "PDFs automatisch durch KI umbenennen.\n"
            "Extrahiert Versicherer, Typ und Datum."
        )
        self.ai_rename_btn.clicked.connect(self._ai_rename_selected)
        header_layout.addWidget(self.ai_rename_btn)
        
        # Upload-Button
        upload_btn = QPushButton("Hochladen")
        upload_btn.clicked.connect(self._upload_document)
        header_layout.addWidget(upload_btn)
        
        layout.addLayout(header_layout)
        
        # Filter-Bereich
        filter_group = QGroupBox("Filter")
        filter_layout = QHBoxLayout(filter_group)
        
        # Quelle-Filter
        filter_layout.addWidget(QLabel("Quelle:"))
        self.source_filter = QComboBox()
        self.source_filter.addItem("Alle", "")
        self.source_filter.addItem("BiPRO (automatisch)", "bipro_auto")
        self.source_filter.addItem("Manuell hochgeladen", "manual_upload")
        self.source_filter.addItem("Selbst erstellt", "self_created")
        self.source_filter.addItem("Scan", "scan")
        self.source_filter.currentIndexChanged.connect(self._apply_filter)
        filter_layout.addWidget(self.source_filter)
        
        # GDV-Filter
        filter_layout.addWidget(QLabel("Typ:"))
        self.type_filter = QComboBox()
        self.type_filter.addItem("Alle", "")
        self.type_filter.addItem("Nur GDV-Dateien", "gdv")
        self.type_filter.currentIndexChanged.connect(self._apply_filter)
        filter_layout.addWidget(self.type_filter)
        
        # Suche
        filter_layout.addWidget(QLabel("Suche:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Dateiname...")
        self.search_input.textChanged.connect(self._filter_table)
        filter_layout.addWidget(self.search_input)
        
        filter_layout.addStretch()
        
        layout.addWidget(filter_group)
        
        # Dokumenten-Tabelle
        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels([
            "ID", "Dateiname", "Quelle", "GDV", "KI", "Groesse", "Hochgeladen", "Von"
        ])
        
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)
        
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)  # Mehrfachauswahl mit Ctrl/Shift
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_context_menu)
        self.table.doubleClicked.connect(self._on_double_click)
        
        layout.addWidget(self.table)
        
        # Status-Zeile
        self.status_label = QLabel("Lade Dokumente...")
        self.status_label.setStyleSheet("color: gray;")
        layout.addWidget(self.status_label)
    
    def refresh_documents(self):
        """Dokumente vom Server laden."""
        self.status_label.setText("Lade Dokumente...")
        self.table.setEnabled(False)
        
        filters = {}
        
        # Quelle-Filter
        source = self.source_filter.currentData()
        if source:
            filters['source'] = source
        
        # Typ-Filter
        type_filter = self.type_filter.currentData()
        if type_filter == 'gdv':
            filters['is_gdv'] = True
        
        self._load_worker = DocumentLoadWorker(self.docs_api, filters)
        self._load_worker.finished.connect(self._on_documents_loaded)
        self._load_worker.error.connect(self._on_load_error)
        self._load_worker.start()
    
    def _on_documents_loaded(self, documents: List[Document]):
        """Callback wenn Dokumente geladen wurden."""
        self._documents = documents
        self._populate_table()
        self.table.setEnabled(True)
        self.status_label.setText(f"{len(documents)} Dokument(e) gefunden")
    
    def _on_load_error(self, error: str):
        """Callback bei Ladefehler."""
        self.table.setEnabled(True)
        self.status_label.setText(f"Fehler: {error}")
        self._toast_manager.show_error(f"Dokumente konnten nicht geladen werden:\n{error}")
    
    def _populate_table(self):
        """Tabelle mit Dokumenten füllen."""
        self.table.setRowCount(len(self._documents))
        
        for row, doc in enumerate(self._documents):
            # ID
            id_item = QTableWidgetItem(str(doc.id))
            id_item.setData(Qt.ItemDataRole.UserRole, doc)
            self.table.setItem(row, 0, id_item)
            
            # Dateiname
            name_item = QTableWidgetItem(doc.original_filename)
            self.table.setItem(row, 1, name_item)
            
            # Quelle
            source_item = QTableWidgetItem(doc.source_type_display)
            if doc.source_type == 'bipro_auto':
                source_item.setForeground(QColor("#2196F3"))
            elif doc.source_type == 'self_created':
                source_item.setForeground(QColor("#4CAF50"))
            elif doc.source_type == 'scan':
                source_item.setForeground(QColor("#9C27B0"))  # Lila fuer Scan
            self.table.setItem(row, 2, source_item)
            
            # GDV
            gdv_item = QTableWidgetItem("Ja" if doc.is_gdv else "")
            if doc.is_gdv:
                gdv_item.setForeground(QColor("#4CAF50"))
            self.table.setItem(row, 3, gdv_item)
            
            # KI-Benennung Status
            if doc.ai_renamed:
                ai_item = QTableWidgetItem("Ja")
                ai_item.setForeground(QColor("#9C27B0"))  # Lila fuer KI
                ai_item.setToolTip("Durch KI umbenannt")
            elif doc.ai_processing_error:
                ai_item = QTableWidgetItem("Fehler")
                ai_item.setForeground(QColor("#F44336"))  # Rot fuer Fehler
                ai_item.setToolTip(f"Fehler: {doc.ai_processing_error}")
            elif doc.is_pdf:
                ai_item = QTableWidgetItem("-")
                ai_item.setToolTip("Noch nicht durch KI verarbeitet")
            else:
                ai_item = QTableWidgetItem("")
                ai_item.setToolTip("Keine PDF-Datei")
            self.table.setItem(row, 4, ai_item)
            
            # Groesse
            size_item = QTableWidgetItem(doc.file_size_display)
            size_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row, 5, size_item)
            
            # Datum im deutschen Format
            date_item = QTableWidgetItem(format_date_german(doc.created_at))
            date_item.setToolTip(doc.created_at or "")  # Original als Tooltip
            self.table.setItem(row, 6, date_item)
            
            # Hochgeladen von
            by_item = QTableWidgetItem(doc.uploaded_by_name or "")
            self.table.setItem(row, 7, by_item)
    
    def _filter_table(self):
        """Tabelle nach Suchbegriff filtern."""
        search_text = self.search_input.text().lower()
        
        for row in range(self.table.rowCount()):
            filename_item = self.table.item(row, 1)
            if filename_item:
                matches = search_text in filename_item.text().lower()
                self.table.setRowHidden(row, not matches)
    
    def _apply_filter(self):
        """Filter anwenden und neu laden."""
        self.refresh_documents()
    
    def _show_context_menu(self, position):
        """Kontextmenü für Tabellenzeilen."""
        item = self.table.itemAt(position)
        if not item:
            return
        
        selected_docs = self._get_selected_documents()
        
        if not selected_docs:
            return
        
        menu = QMenu(self)
        
        if len(selected_docs) == 1:
            # Einzelauswahl
            doc = selected_docs[0]
            
            # PDF Vorschau (als erstes fuer PDFs)
            if self._is_pdf(doc):
                preview_action = QAction("Vorschau", self)
                preview_action.triggered.connect(lambda: self._preview_document(doc))
                menu.addAction(preview_action)
            
            # Download
            download_action = QAction("Herunterladen", self)
            download_action.triggered.connect(lambda: self._download_document(doc))
            menu.addAction(download_action)
            
            # GDV oeffnen
            if doc.is_gdv:
                open_action = QAction("Im GDV-Editor oeffnen", self)
                open_action.triggered.connect(lambda: self._open_in_gdv_editor(doc))
                menu.addAction(open_action)
            
            # KI-Benennung (nur fuer PDFs, die noch nicht umbenannt sind)
            if doc.is_pdf and not doc.ai_renamed:
                menu.addSeparator()
                ai_rename_action = QAction("KI-Benennung", self)
                ai_rename_action.triggered.connect(lambda: self._ai_rename_documents([doc]))
                menu.addAction(ai_rename_action)
            
            menu.addSeparator()
            
            # Loeschen
            delete_action = QAction("Loeschen", self)
            delete_action.triggered.connect(lambda: self._delete_document(doc))
            menu.addAction(delete_action)
        else:
            # Mehrfachauswahl
            download_all_action = QAction(f"{len(selected_docs)} Dokumente herunterladen", self)
            download_all_action.triggered.connect(self._download_selected)
            menu.addAction(download_all_action)
            
            # KI-Benennung fuer mehrere (nur PDFs zaehlen)
            pdf_docs = [d for d in selected_docs if d.is_pdf and not d.ai_renamed]
            if pdf_docs:
                ai_rename_action = QAction(f"KI-Benennung ({len(pdf_docs)} PDFs)", self)
                ai_rename_action.triggered.connect(lambda: self._ai_rename_documents(pdf_docs))
                menu.addAction(ai_rename_action)
            
            menu.addSeparator()
            
            # Mehrfach loeschen
            delete_all_action = QAction(f"{len(selected_docs)} Dokumente loeschen", self)
            delete_all_action.triggered.connect(self._delete_selected)
            menu.addAction(delete_all_action)
        
        menu.exec(self.table.viewport().mapToGlobal(position))
    
    def _on_double_click(self, index):
        """Doppelklick auf Zeile."""
        row = index.row()
        doc_item = self.table.item(row, 0)
        if doc_item:
            doc: Document = doc_item.data(Qt.ItemDataRole.UserRole)
            if doc.is_gdv:
                self._open_in_gdv_editor(doc)
            elif self._is_pdf(doc):
                self._preview_document(doc)
            else:
                self._download_document(doc)
    
    def _is_pdf(self, doc: Document) -> bool:
        """Prüft ob das Dokument ein PDF ist."""
        filename = doc.original_filename.lower()
        mime = (doc.mime_type or "").lower()
        return filename.endswith('.pdf') or 'pdf' in mime
    
    def _preview_selected(self):
        """Zeigt Vorschau für das ausgewählte Dokument."""
        selected_docs = self._get_selected_documents()
        
        if not selected_docs:
            self._toast_manager.show_info("Bitte ein Dokument auswählen.")
            return
        
        if len(selected_docs) > 1:
            self._toast_manager.show_info("Bitte nur ein Dokument für die Vorschau auswählen.")
            return
        
        doc = selected_docs[0]
        
        if self._is_pdf(doc):
            self._preview_document(doc)
        elif doc.is_gdv:
            self._open_in_gdv_editor(doc)
        else:
            self._toast_manager.show_info(
                f"Für '{doc.original_filename}' ist keine Vorschau verfügbar. "
                "Vorschau ist nur für PDF-Dateien und GDV-Dateien möglich."
            )
    
    def _preview_document(self, doc: Document):
        """PDF-Vorschau anzeigen."""
        # Temporäres Verzeichnis
        temp_dir = tempfile.mkdtemp(prefix='bipro_preview_')
        temp_path = os.path.join(temp_dir, doc.original_filename)
        
        # PDF herunterladen
        progress = QProgressDialog("Lade Vorschau...", "Abbrechen", 0, 0, self)
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.show()
        
        try:
            result = self.docs_api.download(doc.id, temp_dir, filename_override=doc.original_filename)
            progress.close()
            
            if result and os.path.exists(result):
                # Viewer öffnen
                viewer = PDFViewerDialog(result, f"Vorschau: {doc.original_filename}", self)
                viewer.exec()
            else:
                self._toast_manager.show_error("PDF konnte nicht geladen werden.")
        except Exception as e:
            progress.close()
            self._toast_manager.show_error(f"Vorschau fehlgeschlagen:\n{e}")
    
    def _upload_document(self):
        """Dokument hochladen."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Dokument hochladen",
            "",
            "Alle Dateien (*);;GDV-Dateien (*.gdv *.txt *.dat);;PDF (*.pdf)"
        )
        
        if not file_path:
            return
        
        # Quelle auswählen
        source_type = 'manual_upload'
        
        # Progress-Dialog
        progress = QProgressDialog("Lade hoch...", "Abbrechen", 0, 0, self)
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.show()
        
        self._upload_worker = UploadWorker(self.docs_api, file_path, source_type)
        self._upload_worker.finished.connect(lambda doc: self._on_upload_finished(doc, progress))
        self._upload_worker.error.connect(lambda err: self._on_upload_error(err, progress))
        self._upload_worker.start()
    
    def _on_upload_finished(self, doc: Optional[Document], progress: QProgressDialog):
        """Callback nach Upload."""
        progress.close()
        
        if doc:
            self._toast_manager.show_success(f"Dokument '{doc.original_filename}' erfolgreich hochgeladen.")
            self.refresh_documents()
        else:
            self._toast_manager.show_error("Upload fehlgeschlagen.")
    
    def _on_upload_error(self, error: str, progress: QProgressDialog):
        """Callback bei Upload-Fehler."""
        progress.close()
        self._toast_manager.show_error(f"Upload fehlgeschlagen:\n{error}")
    
    def _download_document(self, doc: Document):
        """Dokument herunterladen."""
        target_dir = QFileDialog.getExistingDirectory(
            self,
            "Speicherort wählen",
            ""
        )
        
        if not target_dir:
            return
        
        result = self.docs_api.download(doc.id, target_dir, filename_override=doc.original_filename)
        
        if result:
            self._toast_manager.show_success(f"Dokument gespeichert:\n{result}")
        else:
            self._toast_manager.show_error("Download fehlgeschlagen.")
    
    def _open_in_gdv_editor(self, doc: Document):
        """GDV-Dokument im Editor öffnen."""
        self.open_gdv_requested.emit(doc.id, doc.original_filename)
    
    def _delete_document(self, doc: Document):
        """Dokument löschen."""
        reply = QMessageBox.question(
            self,
            "Löschen bestätigen",
            f"Dokument '{doc.original_filename}' wirklich löschen?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            if self.docs_api.delete(doc.id):
                # Erfolgreich gelöscht - keine Meldung, nur Refresh
                self.refresh_documents()
            else:
                # Nur bei Fehler eine Meldung anzeigen
                self._toast_manager.show_error("Löschen fehlgeschlagen.")
    
    def _delete_selected(self):
        """Mehrere ausgewählte Dokumente löschen."""
        selected_docs = self._get_selected_documents()
        
        if not selected_docs:
            return
        
        reply = QMessageBox.question(
            self,
            "Löschen bestätigen",
            f"Wirklich {len(selected_docs)} Dokument(e) löschen?\n\n"
            "Diese Aktion kann nicht rückgängig gemacht werden!",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # Progress-Dialog mit Fortschrittsanzeige
        progress = QProgressDialog(
            "Lösche Dokumente...",
            "Abbrechen",
            0, len(selected_docs),
            self
        )
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setAutoClose(True)  # Automatisch schließen wenn fertig
        progress.setMinimumDuration(0)  # Sofort anzeigen
        
        success_count = 0
        for i, doc in enumerate(selected_docs):
            # Abbruch prüfen
            if progress.wasCanceled():
                break
            
            # Fortschritt aktualisieren
            progress.setValue(i)
            progress.setLabelText(f"Lösche {i+1}/{len(selected_docs)}: {doc.original_filename}")
            QApplication.processEvents()  # UI aktualisieren
            
            # Dokument löschen
            if self.docs_api.delete(doc.id):
                success_count += 1
        
        # Abschluss
        progress.setValue(len(selected_docs))
        
        # Daten neu laden (kein Pop-up nach Abschluss!)
        self.refresh_documents()
    
    def _get_selected_documents(self) -> List[Document]:
        """Gibt alle ausgewählten Dokumente zurück."""
        selected_docs = []
        selected_rows = set()
        
        for item in self.table.selectedItems():
            selected_rows.add(item.row())
        
        for row in selected_rows:
            doc_item = self.table.item(row, 0)
            if doc_item:
                doc = doc_item.data(Qt.ItemDataRole.UserRole)
                if doc:
                    selected_docs.append(doc)
        
        return selected_docs
    
    def _download_selected(self):
        """Ausgewählte Dokumente herunterladen."""
        selected_docs = self._get_selected_documents()
        
        if not selected_docs:
            self._toast_manager.show_info(
                "Bitte mindestens ein Dokument auswählen. "
                "Tipp: Mit Strg+Klick oder Shift+Klick mehrere auswählen."
            )
            return
        
        # Zielordner wählen
        target_dir = QFileDialog.getExistingDirectory(
            self,
            f"Speicherort für {len(selected_docs)} Dokument(e) wählen",
            ""
        )
        
        if not target_dir:
            return
        
        # Downloads durchführen
        success_count = 0
        failed_count = 0
        
        progress = QProgressDialog(
            f"Lade {len(selected_docs)} Dokument(e) herunter...",
            "Abbrechen",
            0, len(selected_docs),
            self
        )
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.show()
        
        for i, doc in enumerate(selected_docs):
            if progress.wasCanceled():
                break
            
            progress.setValue(i)
            progress.setLabelText(f"Lade: {doc.original_filename}")
            
            result = self.docs_api.download(doc.id, target_dir, filename_override=doc.original_filename)
            if result:
                success_count += 1
            else:
                failed_count += 1
        
        progress.close()
        
        # Zusammenfassung
        if failed_count == 0:
            self._toast_manager.show_success(
                f"{success_count} Dokument(e) erfolgreich heruntergeladen. Speicherort: {target_dir}"
            )
        else:
            self._toast_manager.show_warning(
                f"Download: {success_count} erfolgreich, {failed_count} fehlgeschlagen. Speicherort: {target_dir}"
            )
    
    # ========================================
    # KI-Benennung
    # ========================================
    
    def _ai_rename_selected(self):
        """KI-Benennung fuer ausgewaehlte Dokumente starten."""
        selected_docs = self._get_selected_documents()
        
        # Nur PDFs filtern, die noch nicht umbenannt sind
        pdf_docs = [d for d in selected_docs if d.is_pdf and not d.ai_renamed]
        
        if not pdf_docs:
            # Wenn nichts ausgewaehlt oder keine PDFs, alle unbennannten anbieten
            all_unrenamed = [d for d in self._documents if d.is_pdf and not d.ai_renamed]
            
            if not all_unrenamed:
                self._toast_manager.show_info(
                    "Keine PDFs ohne KI-Benennung gefunden. Alle PDFs wurden bereits verarbeitet."
                )
                return
            
            reply = QMessageBox.question(
                self,
                "KI-Benennung",
                f"Keine PDFs ausgewaehlt.\n\n"
                f"Sollen alle {len(all_unrenamed)} unbennannten PDFs verarbeitet werden?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                pdf_docs = all_unrenamed
            else:
                return
        
        self._ai_rename_documents(pdf_docs)
    
    def _ai_rename_documents(self, documents: List[Document]):
        """Startet die KI-Benennung fuer die uebergebenen Dokumente."""
        if not documents:
            return
        
        # Bestaetigung
        reply = QMessageBox.question(
            self,
            "KI-Benennung starten",
            f"{len(documents)} PDF(s) werden durch KI analysiert und umbenannt.\n\n"
            "Das kann je nach Dokumentenanzahl einige Minuten dauern.\n"
            "Die Dokumente werden im Format 'Versicherer_Typ_Datum.pdf' benannt.\n\n"
            "Fortfahren?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # Progress-Dialog
        self._ai_progress = QProgressDialog(
            "Initialisiere KI-Benennung...",
            "Abbrechen",
            0, len(documents),
            self
        )
        self._ai_progress.setWindowTitle("KI-Benennung")
        self._ai_progress.setWindowModality(Qt.WindowModality.WindowModal)
        self._ai_progress.setMinimumDuration(0)
        self._ai_progress.setValue(0)
        self._ai_progress.canceled.connect(self._cancel_ai_rename)
        self._ai_progress.show()
        
        # Worker starten
        self._ai_rename_worker = AIRenameWorker(
            self.api_client,
            self.docs_api,
            documents
        )
        self._ai_rename_worker.progress.connect(self._on_ai_rename_progress)
        self._ai_rename_worker.single_finished.connect(self._on_ai_rename_single)
        self._ai_rename_worker.finished.connect(self._on_ai_rename_finished)
        self._ai_rename_worker.error.connect(self._on_ai_rename_error)
        self._ai_rename_worker.start()
    
    def _cancel_ai_rename(self):
        """Bricht die KI-Benennung ab."""
        if self._ai_rename_worker:
            self._ai_rename_worker.cancel()
    
    def _on_ai_rename_progress(self, current: int, total: int, filename: str):
        """Callback fuer Fortschritt."""
        if hasattr(self, '_ai_progress') and self._ai_progress:
            self._ai_progress.setValue(current)
            self._ai_progress.setLabelText(f"Verarbeite: {filename}\n({current}/{total})")
    
    def _on_ai_rename_single(self, doc_id: int, success: bool, result: str):
        """Callback wenn ein einzelnes Dokument fertig ist."""
        logger.info(f"KI-Benennung Dokument {doc_id}: {'OK' if success else 'FEHLER'} - {result}")
    
    def _on_ai_rename_finished(self, results: List[Tuple[int, bool, str]]):
        """Callback wenn alle Dokumente verarbeitet wurden."""
        if hasattr(self, '_ai_progress') and self._ai_progress:
            self._ai_progress.close()
        
        # Statistik
        success_count = sum(1 for _, success, _ in results if success)
        failed_count = len(results) - success_count
        
        # Detaillierte Ergebnisse
        details = []
        for doc_id, success, result in results[:10]:  # Maximal 10 anzeigen
            status = "OK" if success else "FEHLER"
            details.append(f"  {status}: {result}")
        
        if len(results) > 10:
            details.append(f"  ... und {len(results) - 10} weitere")
        
        detail_text = "\n".join(details)
        
        if failed_count == 0:
            self._toast_manager.show_success(
                f"KI-Benennung: Alle {success_count} Dokument(e) erfolgreich umbenannt."
            )
        else:
            self._toast_manager.show_warning(
                f"KI-Benennung: {success_count} erfolgreich, {failed_count} fehlgeschlagen."
            )
        
        # Tabelle aktualisieren
        self.refresh_documents()
    
    def _on_ai_rename_error(self, error: str):
        """Callback bei globalem Fehler."""
        if hasattr(self, '_ai_progress') and self._ai_progress:
            self._ai_progress.close()
        
        self._toast_manager.show_error(
            f"KI-Benennung Fehler: {error}"
        )


# =============================================================================
# DuplicateCompareDialog - Side-by-Side Vergleich von Duplikaten
# =============================================================================

class DuplicateCompareDialog(QDialog):
    """
    Dialog zum Side-by-Side-Vergleich zweier Duplikat-Dokumente.
    
    Zeigt beide Dokumente nebeneinander mit PDF-Vorschau und
    Aktions-Buttons (Loeschen, Archivieren, Verschieben, Farbe) pro Seite.
    """
    
    documents_changed = Signal()
    
    # Box-Emojis (gleich wie SearchResultCard)
    _BOX_EMOJIS = {
        'gdv': '\U0001f4ca', 'courtage': '\U0001f4b0', 'sach': '\U0001f3e0',
        'leben': '\u2764\ufe0f', 'kranken': '\U0001f3e5', 'sonstige': '\U0001f4c1',
        'roh': '\U0001f4e6', 'eingang': '\U0001f4ec', 'verarbeitung': '\U0001f4e5',
        'falsch': '\u26a0\ufe0f'
    }
    
    # Archivierbare Boxen
    _ARCHIVABLE_BOXES = {'gdv', 'courtage', 'sach', 'leben', 'kranken', 'sonstige'}
    
    def __init__(self, doc_left: Document, doc_right: Document,
                 docs_api: 'DocumentsAPI', preview_cache_dir: str = None, parent=None):
        super().__init__(parent)
        self._doc_left = doc_left
        self._doc_right = doc_right
        self._docs_api = docs_api
        self._preview_cache_dir = preview_cache_dir or os.path.join(
            tempfile.gettempdir(), 'bipro_preview_cache')
        self._has_changes = False
        self._left_disabled = False
        self._right_disabled = False
        self._workers = []
        
        # PDF-Dokument-Objekte fuer QPdfView
        self._pdf_doc_left = None
        self._pdf_doc_right = None
        
        # Direkte Widget-Referenzen (statt findChild)
        self._pdf_views = {}      # side -> QPdfView
        self._loading_labels = {} # side -> QLabel
        
        from i18n.de import DUPLICATE_COMPARE_TITLE
        self.setWindowTitle(DUPLICATE_COMPARE_TITLE)
        self.setMinimumSize(1200, 700)
        self.resize(1400, 900)
        
        self._setup_ui()
        # Previews erst nach show() starten (QPdfView braucht sichtbares Fenster)
        QTimer.singleShot(100, self._download_previews)
    
    def _setup_ui(self):
        """Baut das Dialog-Layout auf."""
        from i18n.de import DUPLICATE_COMPARE_CLOSE
        
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # Splitter fuer Links/Rechts
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setHandleWidth(2)
        
        # Bestimme Labels: Welches ist das "Original"?
        if self._doc_left.is_duplicate:
            left_label_key = 'this'
            right_label_key = 'original'
        elif self._doc_left.is_content_duplicate:
            left_label_key = 'this'
            right_label_key = 'original'
        else:
            left_label_key = 'original'
            right_label_key = 'copy'
        
        # Linke Seite
        self._left_pane = self._build_document_pane(
            self._doc_left, 'left', left_label_key)
        splitter.addWidget(self._left_pane)
        
        # Rechte Seite
        self._right_pane = self._build_document_pane(
            self._doc_right, 'right', right_label_key)
        splitter.addWidget(self._right_pane)
        
        splitter.setSizes([700, 700])
        main_layout.addWidget(splitter, 1)
        
        # Footer mit Schliessen-Button
        footer = QFrame()
        footer.setStyleSheet("QFrame { background: #f5f5f5; border-top: 1px solid #e0e0e0; }")
        footer_layout = QHBoxLayout(footer)
        footer_layout.setContentsMargins(16, 8, 16, 8)
        
        footer_layout.addStretch()
        close_btn = QPushButton(DUPLICATE_COMPARE_CLOSE)
        close_btn.setFixedWidth(140)
        close_btn.setStyleSheet("""
            QPushButton {
                background: #001f3d; color: white; border: none;
                padding: 8px 16px; border-radius: 4px; font-weight: bold;
            }
            QPushButton:hover { background: #002d5c; }
        """)
        close_btn.clicked.connect(self.close)
        footer_layout.addWidget(close_btn)
        footer_layout.addStretch()
        
        main_layout.addWidget(footer)
    
    def _build_document_pane(self, doc: Document, side: str, label_key: str) -> QFrame:
        """Erstellt eine Seite des Vergleichs (Header + Preview + Aktionen)."""
        from i18n.de import (
            DUPLICATE_COMPARE_THIS_DOC, DUPLICATE_COMPARE_COUNTERPART,
            DUPLICATE_COMPARE_COUNTERPART_OF_COPY, DUPLICATE_COMPARE_NO_PREVIEW,
            DUPLICATE_COMPARE_LOADING, DUPLICATE_COMPARE_ACTION_DELETE,
            DUPLICATE_COMPARE_ACTION_ARCHIVE, DUPLICATE_COMPARE_ACTION_UNARCHIVE,
            DUPLICATE_COMPARE_ACTION_MOVE, DUPLICATE_COMPARE_ACTION_COLOR,
            DUPLICATE_TOOLTIP_ARCHIVED
        )
        from api.documents import BOX_DISPLAY_NAMES
        from html import escape
        
        pane = QFrame()
        pane.setObjectName(f"pane_{side}")
        border_color = "#fa9939" if side == 'left' else "#3b82f6"
        pane.setStyleSheet(f"""
            QFrame#{pane.objectName()} {{
                background: white;
                border-top: 3px solid {border_color};
            }}
        """)
        
        layout = QVBoxLayout(pane)
        layout.setContentsMargins(12, 12, 12, 8)
        layout.setSpacing(6)
        
        # --- Header ---
        if label_key == 'this':
            section_label = DUPLICATE_COMPARE_THIS_DOC
        elif label_key == 'original':
            section_label = DUPLICATE_COMPARE_COUNTERPART
        else:
            section_label = DUPLICATE_COMPARE_COUNTERPART_OF_COPY
        
        header_label = QLabel(section_label)
        header_label.setStyleSheet(
            "font-size: 11px; color: #9E9E9E; font-weight: bold; text-transform: uppercase;")
        layout.addWidget(header_label)
        
        # Dateiname
        name_label = QLabel(escape(doc.original_filename))
        name_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #001f3d;")
        name_label.setWordWrap(True)
        layout.addWidget(name_label)
        
        # Meta-Zeile: Box | Datum | Groesse | ggf. Archiviert
        box_emoji = self._BOX_EMOJIS.get(doc.box_type, '\U0001f4c1')
        box_name = BOX_DISPLAY_NAMES.get(doc.box_type, doc.box_type or '')
        date_display = format_date_german(doc.created_at)
        
        meta_parts = [f"{box_emoji} {escape(box_name)}"]
        if date_display:
            meta_parts.append(date_display)
        if doc.file_size:
            size_kb = doc.file_size / 1024
            if size_kb >= 1024:
                meta_parts.append(f"{size_kb / 1024:.1f} MB")
            else:
                meta_parts.append(f"{size_kb:.0f} KB")
        if doc.is_archived:
            meta_parts.append(f"\U0001f4e6 {DUPLICATE_TOOLTIP_ARCHIVED}")
        
        meta_label = QLabel(" | ".join(meta_parts))
        meta_label.setStyleSheet("font-size: 11px; color: #757575;")
        layout.addWidget(meta_label)
        
        # ID
        id_label = QLabel(f"ID: {doc.id}")
        id_label.setStyleSheet("font-size: 10px; color: #BDBDBD;")
        layout.addWidget(id_label)
        
        # --- Vorschau-Bereich ---
        preview_container = QFrame()
        preview_container.setStyleSheet(
            "QFrame { background: #fafafa; border: 1px solid #e0e0e0; border-radius: 4px; }")
        preview_layout = QVBoxLayout(preview_container)
        preview_layout.setContentsMargins(0, 0, 0, 0)
        
        if self._is_pdf(doc) and HAS_PDF_VIEW:
            # Stacked: Loading-Label wird durch QPdfView ersetzt
            from PySide6.QtWidgets import QStackedWidget
            stack = QStackedWidget()
            
            # Seite 0: Loading
            loading_label = QLabel(DUPLICATE_COMPARE_LOADING)
            loading_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            loading_label.setStyleSheet("color: #9E9E9E; font-size: 12px; padding: 40px;")
            stack.addWidget(loading_label)
            
            # Seite 1: QPdfView (sichtbar und gelayoutet von Anfang an)
            pdf_view = QPdfView(stack)
            pdf_view.setPageMode(QPdfView.PageMode.MultiPage)
            pdf_view.setZoomMode(QPdfView.ZoomMode.FitToWidth)
            stack.addWidget(pdf_view)
            
            # Loading-Seite zuerst anzeigen
            stack.setCurrentIndex(0)
            preview_layout.addWidget(stack)
            
            # Direkte Referenzen speichern
            self._pdf_views[side] = pdf_view
            self._loading_labels[side] = stack  # Stack statt Label, um umzuschalten
        else:
            # Kein PDF oder kein QPdfView
            no_preview = QLabel(DUPLICATE_COMPARE_NO_PREVIEW)
            no_preview.setAlignment(Qt.AlignmentFlag.AlignCenter)
            no_preview.setStyleSheet(
                "color: #9E9E9E; font-size: 13px; padding: 60px; font-style: italic;")
            preview_layout.addWidget(no_preview)
        
        layout.addWidget(preview_container, 1)
        
        # --- Status-Overlay (zunachst versteckt) ---
        status_label = QLabel("")
        status_label.setObjectName(f"status_{side}")
        status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        status_label.setStyleSheet("""
            font-size: 14px; font-weight: bold; color: #ef4444;
            padding: 8px; background: #fff5f5; border: 1px solid #fecaca;
            border-radius: 4px;
        """)
        status_label.setVisible(False)
        layout.addWidget(status_label)
        
        # --- Aktions-Buttons ---
        actions_frame = QFrame()
        actions_frame.setObjectName(f"actions_{side}")
        actions_layout = QHBoxLayout(actions_frame)
        actions_layout.setContentsMargins(0, 4, 0, 0)
        actions_layout.setSpacing(6)
        
        # Loeschen
        delete_btn = QPushButton(DUPLICATE_COMPARE_ACTION_DELETE)
        delete_btn.setStyleSheet("""
            QPushButton {
                background: #fee2e2; color: #991b1b; border: 1px solid #fecaca;
                padding: 6px 12px; border-radius: 3px; font-size: 11px;
            }
            QPushButton:hover { background: #fecaca; }
            QPushButton:disabled { background: #f5f5f5; color: #ccc; border: 1px solid #e0e0e0; }
        """)
        delete_btn.clicked.connect(lambda: self._delete_document(side))
        actions_layout.addWidget(delete_btn)
        
        # Archivieren / Entarchivieren
        if doc.box_type in self._ARCHIVABLE_BOXES:
            if doc.is_archived:
                archive_btn = QPushButton(DUPLICATE_COMPARE_ACTION_UNARCHIVE)
                archive_btn.clicked.connect(lambda: self._unarchive_document(side))
            else:
                archive_btn = QPushButton(DUPLICATE_COMPARE_ACTION_ARCHIVE)
                archive_btn.clicked.connect(lambda: self._archive_document(side))
            archive_btn.setStyleSheet("""
                QPushButton {
                    background: #fef3c7; color: #92400e; border: 1px solid #fde68a;
                    padding: 6px 12px; border-radius: 3px; font-size: 11px;
                }
                QPushButton:hover { background: #fde68a; }
                QPushButton:disabled { background: #f5f5f5; color: #ccc; border: 1px solid #e0e0e0; }
            """)
            actions_layout.addWidget(archive_btn)
        
        # Verschieben (mit Dropdown-Menue)
        move_btn = QPushButton(DUPLICATE_COMPARE_ACTION_MOVE)
        move_btn.setStyleSheet("""
            QPushButton {
                background: #e0f2fe; color: #075985; border: 1px solid #bae6fd;
                padding: 6px 12px; border-radius: 3px; font-size: 11px;
            }
            QPushButton:hover { background: #bae6fd; }
            QPushButton::menu-indicator { subcontrol-position: right center; }
            QPushButton:disabled { background: #f5f5f5; color: #ccc; border: 1px solid #e0e0e0; }
        """)
        move_menu = QMenu(move_btn)
        move_targets = ['gdv', 'courtage', 'sach', 'leben', 'kranken', 'sonstige', 'eingang', 'roh']
        for box_type in move_targets:
            if box_type == doc.box_type:
                continue  # Aktuelle Box ueberspringen
            emoji = self._BOX_EMOJIS.get(box_type, '\U0001f4c1')
            display = BOX_DISPLAY_NAMES.get(box_type, box_type)
            action = move_menu.addAction(f"{emoji} {display}")
            action.triggered.connect(
                lambda checked, bt=box_type, s=side: self._move_document(s, bt))
        move_btn.setMenu(move_menu)
        actions_layout.addWidget(move_btn)
        
        # Farbe (mit Dropdown-Menue)
        color_btn = QPushButton(DUPLICATE_COMPARE_ACTION_COLOR)
        color_btn.setStyleSheet("""
            QPushButton {
                background: #f3e8ff; color: #6b21a8; border: 1px solid #e9d5ff;
                padding: 6px 12px; border-radius: 3px; font-size: 11px;
            }
            QPushButton:hover { background: #e9d5ff; }
            QPushButton::menu-indicator { subcontrol-position: right center; }
            QPushButton:disabled { background: #f5f5f5; color: #ccc; border: 1px solid #e0e0e0; }
        """)
        from ui.styles.tokens import DOCUMENT_DISPLAY_COLORS
        from i18n.de import (DOC_COLOR_GREEN, DOC_COLOR_RED, DOC_COLOR_BLUE,
                              DOC_COLOR_ORANGE, DOC_COLOR_PURPLE, DOC_COLOR_PINK,
                              DOC_COLOR_CYAN, DOC_COLOR_YELLOW, DOC_COLOR_REMOVE)
        color_labels = {
            'green': DOC_COLOR_GREEN, 'red': DOC_COLOR_RED, 'blue': DOC_COLOR_BLUE,
            'orange': DOC_COLOR_ORANGE, 'purple': DOC_COLOR_PURPLE, 'pink': DOC_COLOR_PINK,
            'cyan': DOC_COLOR_CYAN, 'yellow': DOC_COLOR_YELLOW,
        }
        color_menu = QMenu(color_btn)
        for color_key, color_label in color_labels.items():
            hex_color = DOCUMENT_DISPLAY_COLORS.get(color_key, '#ccc')
            action = color_menu.addAction(f"\u25cf {color_label}")
            action.triggered.connect(
                lambda checked, ck=color_key, s=side: self._color_document(s, ck))
        color_menu.addSeparator()
        remove_action = color_menu.addAction(DOC_COLOR_REMOVE)
        remove_action.triggered.connect(lambda: self._color_document(side, None))
        color_btn.setMenu(color_menu)
        actions_layout.addWidget(color_btn)
        
        actions_layout.addStretch()
        layout.addWidget(actions_frame)
        
        return pane
    
    def _is_pdf(self, doc: Document) -> bool:
        """Prueft ob ein Dokument ein PDF ist."""
        if doc.mime_type and 'pdf' in doc.mime_type.lower():
            return True
        name = (doc.original_filename or '').lower()
        return name.endswith('.pdf')
    
    def _download_previews(self):
        """Startet den Download beider PDF-Vorschauen parallel."""
        os.makedirs(self._preview_cache_dir, exist_ok=True)
        
        for side, doc in [('left', self._doc_left), ('right', self._doc_right)]:
            if not self._is_pdf(doc) or not HAS_PDF_VIEW:
                continue
            
            # Cache-Check (sanitisierter Dateiname fuer Windows-Kompatibilitaet)
            from api.documents import safe_cache_filename
            cached = os.path.join(self._preview_cache_dir,
                                  safe_cache_filename(doc.id, doc.original_filename))
            if os.path.exists(cached) and os.path.getsize(cached) > 0:
                self._on_preview_ready(side, cached)
                continue
            
            # Download starten
            from ui.archive_boxes_view import PreviewDownloadWorker
            worker = PreviewDownloadWorker(
                self._docs_api, doc.id, self._preview_cache_dir,
                filename=doc.original_filename,
                cache_dir=self._preview_cache_dir)
            worker.download_finished.connect(
                lambda path, s=side: self._on_preview_ready(s, path))
            worker.download_error.connect(
                lambda err, s=side: self._on_preview_error(s, err))
            self._workers.append(worker)
            worker.start()
    
    def _on_preview_ready(self, side: str, path):
        """Callback wenn PDF-Download fertig ist."""
        if not path or not os.path.exists(path):
            self._on_preview_error(side, "Datei nicht gefunden")
            return
        
        pdf_view = self._pdf_views.get(side)
        stack = self._loading_labels.get(side)
        
        if pdf_view:
            pdf_doc = QPdfDocument(self)
            
            # QPdfDocument.load() kann fehlschlagen bei Sonderzeichen im Pfad.
            # Wir pruefen den Status und laden ggf. von QBuffer als Fallback.
            load_error = pdf_doc.load(path)
            
            # Status pruefen (Error = 2 in QPdfDocument.Error enum)
            if pdf_doc.status() == QPdfDocument.Status.Error:
                logger.warning(f"QPdfDocument.load() fehlgeschlagen fuer {path}, versuche QBuffer-Fallback")
                try:
                    from PySide6.QtCore import QBuffer, QByteArray
                    with open(path, 'rb') as f:
                        data = f.read()
                    # Neues QPdfDocument fuer Buffer-Modus
                    pdf_doc = QPdfDocument(self)
                    self._pdf_buffer = QBuffer(self)  # Buffer muss am Leben bleiben
                    self._pdf_buffer.setData(QByteArray(data))
                    self._pdf_buffer.open(QBuffer.OpenModeFlag.ReadOnly)
                    pdf_doc.load(self._pdf_buffer)
                except Exception as e:
                    logger.error(f"QBuffer-Fallback fehlgeschlagen ({side}): {e}")
                    self._on_preview_error(side, str(e))
                    return
            
            pdf_view.setDocument(pdf_doc)
            
            # Stack auf QPdfView-Seite (Index 1) umschalten
            if stack:
                stack.setCurrentIndex(1)
            
            if side == 'left':
                self._pdf_doc_left = pdf_doc
            else:
                self._pdf_doc_right = pdf_doc
            
            logger.info(f"PDF-Vorschau geladen ({side}): {path}")
    
    def _on_preview_error(self, side: str, error: str):
        """Callback bei Download-Fehler."""
        stack = self._loading_labels.get(side)
        if stack:
            # Loading-Label (Index 0) mit Fehlermeldung aktualisieren
            loading = stack.widget(0)
            if loading:
                from i18n.de import DUPLICATE_COMPARE_NO_PREVIEW
                loading.setText(DUPLICATE_COMPARE_NO_PREVIEW)
                loading.setStyleSheet(
                    "color: #ef4444; font-size: 12px; padding: 40px; font-style: italic;")
        logger.warning(f"PDF-Vorschau Fehler ({side}): {error}")
    
    def _get_doc(self, side: str) -> Document:
        """Gibt das Dokument fuer die angegebene Seite zurueck."""
        return self._doc_left if side == 'left' else self._doc_right
    
    def _is_side_disabled(self, side: str) -> bool:
        """Prueft ob eine Seite bereits deaktiviert ist."""
        return self._left_disabled if side == 'left' else self._right_disabled
    
    def _mark_pane_modified(self, side: str, status_text: str):
        """Markiert eine Seite als modifiziert (deaktiviert Buttons, zeigt Status)."""
        pane = self._left_pane if side == 'left' else self._right_pane
        
        # Status-Label anzeigen
        status = pane.findChild(QLabel, f"status_{side}")
        if status:
            status.setText(status_text)
            status.setVisible(True)
        
        # Aktions-Buttons deaktivieren
        actions = pane.findChild(QFrame, f"actions_{side}")
        if actions:
            for btn in actions.findChildren(QPushButton):
                btn.setEnabled(False)
        
        # Seite als deaktiviert markieren
        if side == 'left':
            self._left_disabled = True
        else:
            self._right_disabled = True
        
        self._has_changes = True
    
    def _delete_document(self, side: str):
        """Loescht das Dokument auf der angegebenen Seite."""
        if self._is_side_disabled(side):
            return
        
        doc = self._get_doc(side)
        from i18n.de import DUPLICATE_COMPARE_CONFIRM_DELETE, DUPLICATE_COMPARE_DELETED
        
        reply = QMessageBox.question(
            self, DUPLICATE_COMPARE_DELETED,
            DUPLICATE_COMPARE_CONFIRM_DELETE.format(filename=doc.original_filename),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        try:
            self._docs_api.delete(doc.id)
            self._mark_pane_modified(side, f"\u274c {DUPLICATE_COMPARE_DELETED}")
        except Exception as e:
            from i18n.de import DUPLICATE_COMPARE_ERROR
            QMessageBox.warning(self, "Fehler",
                                DUPLICATE_COMPARE_ERROR.format(error=str(e)))
    
    def _archive_document(self, side: str):
        """Archiviert das Dokument."""
        if self._is_side_disabled(side):
            return
        doc = self._get_doc(side)
        try:
            self._docs_api.archive_documents([doc.id])
            from i18n.de import DUPLICATE_COMPARE_ARCHIVED
            self._mark_pane_modified(side, f"\U0001f4e6 {DUPLICATE_COMPARE_ARCHIVED}")
        except Exception as e:
            from i18n.de import DUPLICATE_COMPARE_ERROR
            QMessageBox.warning(self, "Fehler",
                                DUPLICATE_COMPARE_ERROR.format(error=str(e)))
    
    def _unarchive_document(self, side: str):
        """Entarchiviert das Dokument."""
        if self._is_side_disabled(side):
            return
        doc = self._get_doc(side)
        try:
            self._docs_api.unarchive_documents([doc.id])
            from i18n.de import DUPLICATE_COMPARE_UNARCHIVED
            self._mark_pane_modified(side, f"\U0001f4e4 {DUPLICATE_COMPARE_UNARCHIVED}")
        except Exception as e:
            from i18n.de import DUPLICATE_COMPARE_ERROR
            QMessageBox.warning(self, "Fehler",
                                DUPLICATE_COMPARE_ERROR.format(error=str(e)))
    
    def _move_document(self, side: str, target_box: str):
        """Verschiebt das Dokument in eine andere Box."""
        if self._is_side_disabled(side):
            return
        doc = self._get_doc(side)
        try:
            self._docs_api.move_documents([doc.id], target_box)
            from i18n.de import DUPLICATE_COMPARE_MOVED
            from api.documents import BOX_DISPLAY_NAMES
            box_name = BOX_DISPLAY_NAMES.get(target_box, target_box)
            emoji = self._BOX_EMOJIS.get(target_box, '\U0001f4c1')
            self._mark_pane_modified(
                side, f"{emoji} {DUPLICATE_COMPARE_MOVED.format(box=box_name)}")
        except Exception as e:
            from i18n.de import DUPLICATE_COMPARE_ERROR
            QMessageBox.warning(self, "Fehler",
                                DUPLICATE_COMPARE_ERROR.format(error=str(e)))
    
    def _color_document(self, side: str, color_key):
        """Setzt die Farbmarkierung des Dokuments."""
        if self._is_side_disabled(side):
            return
        doc = self._get_doc(side)
        try:
            self._docs_api.set_documents_color([doc.id], color_key)
            from i18n.de import DUPLICATE_COMPARE_COLORED, DUPLICATE_COMPARE_COLOR_REMOVED
            if color_key:
                self._mark_pane_modified(side, f"\U0001f3a8 {DUPLICATE_COMPARE_COLORED}")
            else:
                self._mark_pane_modified(side, DUPLICATE_COMPARE_COLOR_REMOVED)
        except Exception as e:
            from i18n.de import DUPLICATE_COMPARE_ERROR
            QMessageBox.warning(self, "Fehler",
                                DUPLICATE_COMPARE_ERROR.format(error=str(e)))
    
    def closeEvent(self, event):
        """Beim Schliessen: Worker stoppen und ggf. Signal senden."""
        for worker in self._workers:
            if worker.isRunning():
                worker.cancel()
                worker.wait(1000)
        
        # PDF-Dokumente freigeben
        if self._pdf_doc_left:
            self._pdf_doc_left.close()
        if self._pdf_doc_right:
            self._pdf_doc_right.close()
        
        if self._has_changes:
            self.documents_changed.emit()
        
        super().closeEvent(event)
