"""
Xempus-Parser fuer alle 5 Sheets.

Parst ALLE Zeilen aus ALLEN Sheets, ueberspringt NICHTS.
Jede Zeile wird als Dict zurueckgegeben (raw_json-faehig).
"""

import hashlib
import json
import logging
from dataclasses import dataclass, field
from datetime import datetime
from typing import List, Dict, Optional, Any

logger = logging.getLogger(__name__)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl nicht installiert -- Excel-Import nicht verfuegbar")


XEMPUS_SHEETS = ['ArbG', 'ArbG-Tarife', 'ArbG-Zuschüsse', 'ArbN', 'Beratungen']


def _col_index(letter: str) -> int:
    """Spaltenbuchstabe in 1-basierten Index umrechnen."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result


def _safe_str(cell) -> Optional[str]:
    """Zellwert sicher als String extrahieren."""
    if cell is None:
        return None
    val = cell.value if hasattr(cell, 'value') else cell
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _safe_float(cell) -> Optional[float]:
    """Zellwert sicher als Float extrahieren."""
    val = cell.value if hasattr(cell, 'value') else cell
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s == '-':
        return None
    s = s.replace(' ', '').replace('€', '').replace('EUR', '')
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def _safe_int(cell) -> Optional[int]:
    val = cell.value if hasattr(cell, 'value') else cell
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        return None


def _safe_bool(cell) -> Optional[bool]:
    val = cell.value if hasattr(cell, 'value') else cell
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ('ja', 'yes', '1', 'true', 'x'):
        return True
    if s in ('nein', 'no', '0', 'false', ''):
        return False
    return None


def _safe_date(cell) -> Optional[str]:
    """Datum als YYYY-MM-DD String extrahieren."""
    val = cell.value if hasattr(cell, 'value') else cell
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s:
        return None
    for fmt in ('%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def _row_to_raw_dict(cells, headers: List[str]) -> Dict[str, Any]:
    """Zeile als Dict mit Header-Keys serialisieren (fuer raw_json)."""
    result = {}
    for i, h in enumerate(headers):
        if i < len(cells):
            val = cells[i].value if hasattr(cells[i], 'value') else cells[i]
            if isinstance(val, datetime):
                val = val.strftime('%Y-%m-%d')
            elif isinstance(val, (int, float)):
                pass
            elif val is not None:
                val = str(val)
            result[h] = val
    return result


ARBG_COLUMNS = {
    'A': 'name', 'B': 'street', 'C': 'plz', 'D': 'city',
    'E': 'iban', 'F': 'bic', 'G': 'tarif_info', 'H': 'zuschuss_info',
    'I': 'id',
}

ARBG_TARIFE_COLUMNS = {
    'B': 'versicherer', 'C': 'typ', 'D': 'durchfuehrungsweg',
    'E': 'tarif', 'F': 'beantragung',
    'G': 'gruppenrahmenkollektiv', 'H': 'gruppennummer',
    'I': 'employer_id', 'J': 'id',
}

ARBG_ZUSCHUSS_COLUMNS = {
    'B': 'bezeichnung', 'C': 'art_vl_umwandlung',
    'D': 'zuschuss_vl_alternativ', 'E': 'prozent_auf_vl',
    'F': 'zuschuss_prozentual_leq_bbg', 'G': 'zuschuss_prozentual_gt_bbg',
    'H': 'begrenzung_prozentual', 'I': 'fester_zuschuss', 'J': 'fester_arbg_beitrag',
    'K': 'gestaffelter_zuschuss_aktiv', 'L': 'gestaffelter_zuschuss',
    'M': 'begrenzung_gestaffelt', 'N': 'employer_id', 'O': 'id',
}

ARBN_COLUMNS = {
    'B': 'anrede', 'C': 'titel', 'D': 'name', 'E': 'vorname',
    'F': 'beratungsstatus', 'G': 'street', 'H': 'plz', 'I': 'city',
    'J': 'bundesland', 'K': 'land', 'L': 'geburtsdatum',
    'M': 'telefon', 'N': 'mobiltelefon', 'O': 'email',
    'P': 'diensteintritt', 'Q': 'krankenversicherung',
    'R': 'bruttolohn', 'S': 'steuerklasse',
    'T': 'berufsstellung', 'U': 'berufsbezeichnung',
    'V': 'personalnummer', 'W': 'staatsangehoerigkeit',
    'X': 'familienstand', 'Y': 'kinder_vorhanden',
    'Z': 'kinderfreibetrag', 'AA': 'freibetrag_jaehrlich',
    'AB': 'kirchensteuerpflicht', 'AC': 'bemerkung',
    'AD': 'zuschuss_name', 'AE': 'id',
    'AF': 'employer_id', 'AG': 'zuschuss_id',
}

BERATUNGEN_COLUMNS = {
    'A': 'arbg_name', 'B': 'arbn_name', 'C': 'arbn_vorname',
    'D': 'geburtsdatum', 'E': 'status', 'F': 'beratungsdatum',
    'G': 'beginn', 'H': 'ende',
    'I': 'arbn_anteil', 'J': 'davon_vl_arbn',
    'K': 'arbg_anteil', 'L': 'davon_vl_arbg',
    'M': 'gesamtbeitrag', 'N': 'entgeltumwandlung_aus',
    'O': 'versicherungsscheinnummer', 'P': 'versicherer',
    'Q': 'typ', 'R': 'durchfuehrungsweg', 'S': 'tarif',
    'T': 'beantragung', 'U': 'tarifoption', 'V': 'gruppennummer',
    'W': 'buz', 'X': 'buz_rente', 'Y': 'dauer_jahre',
    'Z': 'garantierte_rente', 'AA': 'garantierte_kapitalleistung',
    'AB': 'sbu_jahresbruttolohn', 'AC': 'sbu_garantierte_bu_rente',
    'AD': 'sbu_gesamte_bu_rente', 'AE': 'rentenalter',
    'AF': 'berater', 'AG': 'beratungstyp', 'AH': 'zahlungsweise',
    'AI': 'agenturnummer',
    'AJ': 'datum_antragsdokument', 'AK': 'datum_entscheidung',
    'AL': 'datum_elektronische_uebermittlung',
    'AM': 'id', 'AN': 'employee_id', 'AO': 'employer_id',
}

SHEET_COLUMN_MAPS = {
    'ArbG': ARBG_COLUMNS,
    'ArbG-Tarife': ARBG_TARIFE_COLUMNS,
    'ArbG-Zuschüsse': ARBG_ZUSCHUSS_COLUMNS,
    'ArbN': ARBN_COLUMNS,
    'Beratungen': BERATUNGEN_COLUMNS,
}

DECIMAL_FIELDS = {
    'zuschuss_vl_alternativ', 'zuschuss_prozentual_leq_bbg', 'zuschuss_prozentual_gt_bbg',
    'fester_zuschuss', 'fester_arbg_beitrag', 'bruttolohn', 'kinderfreibetrag',
    'freibetrag_jaehrlich', 'arbn_anteil', 'davon_vl_arbn', 'arbg_anteil',
    'davon_vl_arbg', 'gesamtbeitrag', 'buz_rente', 'garantierte_rente',
    'garantierte_kapitalleistung', 'sbu_jahresbruttolohn',
    'sbu_garantierte_bu_rente', 'sbu_gesamte_bu_rente',
}

DATE_FIELDS = {
    'geburtsdatum', 'diensteintritt', 'beratungsdatum', 'beginn', 'ende',
    'datum_antragsdokument', 'datum_entscheidung', 'datum_elektronische_uebermittlung',
}

BOOL_FIELDS = {
    'prozent_auf_vl', 'gestaffelter_zuschuss_aktiv', 'kinder_vorhanden',
    'kirchensteuerpflicht', 'buz',
}

INT_FIELDS = {'dauer_jahre', 'rentenalter'}


def _parse_cell_by_field(cell, field_name: str) -> Any:
    """Zellwert anhand des Feldnamens typgerecht parsen."""
    if field_name in DECIMAL_FIELDS:
        return _safe_float(cell)
    if field_name in DATE_FIELDS:
        return _safe_date(cell)
    if field_name in BOOL_FIELDS:
        return _safe_bool(cell)
    if field_name in INT_FIELDS:
        return _safe_int(cell)
    return _safe_str(cell)


@dataclass
class SheetParseResult:
    """Ergebnis eines einzelnen Sheet-Parsings."""
    sheet_name: str = ''
    rows: List[Dict] = field(default_factory=list)
    total_rows: int = 0
    errors: List[Dict] = field(default_factory=list)


@dataclass
class XempusCompleteResult:
    """Ergebnis des vollstaendigen Xempus-Parsings (alle 5 Sheets)."""
    sheets: Dict[str, SheetParseResult] = field(default_factory=dict)
    filename: str = ''
    sheets_found: List[str] = field(default_factory=list)
    total_rows: int = 0
    total_errors: int = 0

    def get_sheet_data(self, sheet_name: str) -> SheetParseResult:
        return self.sheets.get(sheet_name, SheetParseResult(sheet_name=sheet_name))


def _detect_headers(ws, max_cols: int = 100) -> List[str]:
    """Header-Zeile (Zeile 1) lesen und als Liste von Strings zurueckgeben."""
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, max_col=max_cols):
        for cell in row:
            headers.append(str(cell.value or '').strip())
        break
    return headers


def _parse_sheet_generic(ws, sheet_name: str, column_map: Dict[str, str]) -> SheetParseResult:
    """Ein Sheet generisch parsen basierend auf Column-Map.

    Ueberspringt NICHTS - jede Zeile wird verarbeitet.
    """
    result = SheetParseResult(sheet_name=sheet_name)

    col_indices = {_col_index(letter): field_name for letter, field_name in column_map.items()}
    max_col_needed = max(col_indices.keys()) if col_indices else 1

    headers = _detect_headers(ws, max_cols=max_col_needed + 10)

    for row in ws.iter_rows(min_row=2, max_col=max_col_needed + 20):
        result.total_rows += 1
        row_num = result.total_rows + 1

        try:
            cells = list(row)

            all_empty = all(
                (c.value is None or str(c.value).strip() == '')
                for c in cells[:max_col_needed]
            )
            if all_empty:
                continue

            parsed = {}
            for col_idx, field_name in col_indices.items():
                cell_idx = col_idx - 1
                if cell_idx < len(cells):
                    parsed[field_name] = _parse_cell_by_field(cells[cell_idx], field_name)
                else:
                    parsed[field_name] = None

            if sheet_name == 'Beratungen':
                extra = {}
                for i, cell in enumerate(cells):
                    if (i + 1) not in col_indices and cell.value is not None:
                        col_header = headers[i] if i < len(headers) else f'Spalte{i+1}'
                        val = cell.value
                        if isinstance(val, datetime):
                            val = val.strftime('%Y-%m-%d')
                        extra[col_header] = val
                if extra:
                    parsed['extra_cols'] = extra

            raw_dict = _row_to_raw_dict(cells, headers)
            parsed['_raw'] = raw_dict

            result.rows.append(parsed)

        except Exception as e:
            result.errors.append({
                'row': row_num,
                'error': str(e),
                'raw': _row_to_raw_dict(list(row), headers) if headers else {},
            })

    return result


def parse_xempus_complete(filepath: str) -> XempusCompleteResult:
    """Parst ALLE 5 Sheets einer Xempus-Excel-Datei.

    Ueberspringt KEINE Zeile. Jede Zeile wird als Dict mit
    raw_json-faehigem _raw-Feld zurueckgegeben.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl nicht installiert")

    import time
    t0 = time.time()

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    result = XempusCompleteResult(
        filename=filepath,
        sheets_found=list(wb.sheetnames),
    )

    for sheet_name, column_map in SHEET_COLUMN_MAPS.items():
        actual_name = None
        for sn in wb.sheetnames:
            if sn == sheet_name or sn.lower() == sheet_name.lower():
                actual_name = sn
                break
        if sheet_name == 'ArbG-Zuschüsse':
            for sn in wb.sheetnames:
                if 'zusch' in sn.lower():
                    actual_name = sn
                    break

        if not actual_name:
            logger.warning(f"Sheet '{sheet_name}' nicht gefunden in {filepath}")
            result.sheets[sheet_name] = SheetParseResult(sheet_name=sheet_name)
            continue

        ws = wb[actual_name]
        sheet_result = _parse_sheet_generic(ws, sheet_name, column_map)
        result.sheets[sheet_name] = sheet_result
        result.total_rows += sheet_result.total_rows
        result.total_errors += len(sheet_result.errors)

        logger.info(
            f"Xempus Sheet '{sheet_name}': {len(sheet_result.rows)} Zeilen, "
            f"{len(sheet_result.errors)} Fehler"
        )

    wb.close()
    elapsed = time.time() - t0
    logger.info(f"Xempus komplett: {result.total_rows} Zeilen in {elapsed:.1f}s")

    return result


def prepare_sheets_for_upload(parse_result: XempusCompleteResult) -> List[Dict]:
    """Konvertiert ParseResult in das Format fuer den POST /pm/xempus/import Endpoint.

    Returns: Liste von {sheet_name, rows: [{...}]} Dicts.
    """
    sheets = []
    for sheet_name, sheet_result in parse_result.sheets.items():
        if not sheet_result.rows and not sheet_result.errors:
            continue

        rows = []
        for row in sheet_result.rows:
            row_copy = dict(row)
            row_copy.pop('_raw', None)
            rows.append(row_copy)

        for err in sheet_result.errors:
            raw = err.get('raw', {})
            raw['_parse_error'] = err.get('error', '')
            rows.append(raw)

        sheets.append({
            'sheet_name': sheet_name,
            'rows': rows,
        })

    return sheets
